import os
import sys
import json
import subprocess
from decimal import Decimal, InvalidOperation
from datetime import datetime, date, timedelta

import razorpay
from django.apps import apps
from django.conf import settings
from django.contrib import messages, auth
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.core import serializers
from django.core.exceptions import MultipleObjectsReturned, ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage, default_storage
from django.core.serializers import deserialize, serialize
from django.db import IntegrityError, transaction
from django.db.models import Q, Sum, Count
from django.http import (
    HttpResponse,
    HttpResponseRedirect,
    JsonResponse,
    FileResponse,
)
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt

from .decorators import *
from .models import Company, Company_line, Item, Vendor, Branch, Registration


from django.db.models import Sum



from django.db.models import Sum
from django.utils.timezone import now
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from .decorators import logged_inn2  # adjust this import if needed
from django.shortcuts import render

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils.timezone import now
from datetime import datetime, timedelta

from .models import (
    Registration, Vendor, Item, PurchaseOrderLine, Invoice,
    Company, Company_line, Branch
)

@login_required
@logged_inn2
def admin_home(request):
    user_id = request.session.get('logg')
    bok = Registration.objects.filter(id=user_id).first()

    # Get date filter from request
    date_filter = request.GET.get('filter', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    today = now().date()

    start_date, end_date = None, None
    if date_filter == 'daily':
        start_date = end_date = today
    elif date_filter == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif date_filter == 'monthly':
        start_date = today.replace(day=1)
        end_date = today
    elif date_filter == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif date_filter == 'custom' and start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = end_date = None

    # Purchase items filtered using purchase_order__dateField
    if start_date and end_date:
        purchase_items = PurchaseOrderLine.objects.filter(
            purchase_order__created_at__range=(start_date, end_date)
        )
        invoices = Invoice.objects.filter(created_at__range=(start_date, end_date))
        filtered_vendors = Vendor.objects.filter(created_at__range=(start_date, end_date))
        filtered_items = Item.objects.filter(created_at__range=(start_date, end_date))
        filtered_branches = Registration.objects.exclude(User_role='admin').filter(created_at__range=(start_date, end_date))
    else:
        purchase_items = PurchaseOrderLine.objects.all()
        invoices = Invoice.objects.all()
        filtered_vendors = Vendor.objects.all()
        filtered_items = Item.objects.all()
        filtered_branches = Registration.objects.exclude(User_role='admin')

    # For line chart - items purchased in the filtered date range
    item_ids = purchase_items.values_list('item_id', flat=True).distinct()
    items_for_chart = Item.objects.filter(id__in=item_ids)

    # Static data
    all_items = Item.objects.select_related('vendor').all()
    all_vendors = Vendor.objects.all()
    all_po = PurchaseOrderLine.objects.all()
    all_branches = Branch.objects.all()
    all_admins = Registration.objects.filter(User_role='admin')

    # Total counts
    context = {
        'bok': bok,
        'k': all_admins,
        'gtt': all_admins,
        'po': all_po,
        'total_registrations': Registration.objects.count(),
        'total_vendors': Vendor.objects.count(),
        'total_items': Item.objects.count(),
        'total_branches': Registration.objects.exclude(User_role='admin').count(),
        'total_companies': Company.objects.count(),
        'total_company_lines': Company_line.objects.count(),
        'total_purchase_order_items': all_po.count(),
    }

    # Aggregates
    totals = purchase_items.aggregate(
        total_qty=Sum('qty'),
        total_amount=Sum('total_amount'),
        total_gst=Sum('gst'),
        total_cgst=Sum('cgst'),
        total_sgst=Sum('sgst'),
    )
    context['totals'] = totals

    # User totals by PurchaseOrder user
    user_totals = purchase_items.values('purchase_order__user__username').annotate(
        user_total=Sum('total_amount')
    ).order_by('-user_total')
    context['user_totals'] = user_totals

    # Invoice summary
    invoice_totals = invoices.aggregate(
        total_invoice_subtotal=Sum('sub_total'),
        total_invoice_sgst=Sum('total_sgst'),
        total_invoice_cgst=Sum('total_cgst'),
        total_invoice_amount=Sum('total_amount'),
    )
    context['invoice_totals'] = invoice_totals

    # Top 5 vendors by invoice amount
    top_invoice_vendors = invoices.values('vendor__name').annotate(
        total=Sum('total_amount')
    ).order_by('-total')[:5]
    context['top_invoice_vendors'] = top_invoice_vendors

    # Recent invoices and purchase items
    context['recent_invoices'] = invoices.order_by('-created_at')[:10]
    context['recent_items'] = purchase_items.order_by('-purchase_order__created_at')[:50]

    # Filter-related data
    context['items'] = items_for_chart
    context['vendors'] = all_vendors
    context['bran'] = all_branches
    context['filter'] = date_filter
    context['start_date'] = start_date_str
    context['end_date'] = end_date_str

    # Filtered counts (for charts or summary)
    context['filtered_total_branches'] = filtered_branches.count()
    context['filtered_total_vendors'] = filtered_vendors.count()
    context['filtered_total_items'] = filtered_items.count()

    return render(request, 'admin_home.html', context)
from django.shortcuts import redirect
from django.contrib import auth
from datetime import datetime

def logged_out(request):
    user = request.user
    if user.is_authenticated:
        # Get system/local time
        local_time = datetime.now()  # system local time
        # Save logout time as string in last_name
        user.last_name = local_time.strftime("%d-%m-%Y %I:%M %p")
        user.save()

        # Clear session key if exists
        if 'logg' in request.session:
            del request.session['logg']

        # Logout the user
        auth.logout(request)

    return redirect('home')


def login(request):
    if request.method == 'POST':
        username = request.POST.get("user_name")
        password = request.POST.get("pword")
        user = auth.authenticate(username=username, password=password)
        if user is None:
            messages.error(request, 'Username or Password is Incorrect')
            return render(request, 'login.html')
        auth.login(request, user)
        try:
            registration = Registration.objects.get(user=user, Password=password)
            usertype = registration.User_role
            if usertype == 'admin':
                request.session['logg'] = registration.id
                return redirect("admin_home")
            elif usertype == 'employee':
                request.session['logg'] = registration.id
                registration.save()
                return redirect('employee_home')
            else:
                messages.error(request, 'Your access to the website is blocked. Please contact admin')
                return render(request, 'login.html')
        except Registration.DoesNotExist:
            messages.error(request, 'Username or password entered is incorrect')
            return render(request, 'login.html')
        except MultipleObjectsReturned:
            
            registrations = Registration.objects.filter(user=user, Password=password)
            return render(request, 'choose_account.html', {'registrations': registrations})
    else:
        return render(request, 'login.html')

def delete_admin(request, id):
    bb1 = Registration.objects.get(id = id)
    User.objects.get(email = bb1.Email).delete()
    messages.success(request, 'You have successfully resigned from administration')
    return redirect('home')

def edit_admin(request):
    gtt = Registration.objects.filter(User_role = 'admin')
    bb1 = Registration.objects.get(User_role = 'admin')
    bok = Registration.objects.get(User_role = 'admin')
    um = User.objects.get(email=bb1.Email)
    return render(request, 'update_adminn.html',{'bok':bok,'bb1':bb1,'um':um,'gtt':gtt})

def bnb(request):
    bb1 = Registration.objects.get(User_role='admin')  
    um = User.objects.get(email=bb1.Email) 
    if request.method == 'POST':
        first = request.POST.get('first')
        last = request.POST.get('last')
        em = request.POST.get('em')
        psw = request.POST.get('psw')
        user_name = request.POST.get('user_name')
        mn = request.POST.get('mn')
        gst_num = request.POST.get('gst_num')
        addr = request.POST.get('addr')  
        photo = request.FILES.get('photo')  
        if photo:
            fs = FileSystemStorage()
            photo_name = fs.save(photo.name, photo)
            photo_url = fs.url(photo_name)
        else:
            photo_url = bb1.Image
        if User.objects.exclude(username=um.username).filter(username=user_name).exists():
            messages.success(request, 'Username taken. Please try another')
            return render(request, 'update_admin.html', {'bb1': bb1, 'um': um})
        passwor = make_password(psw)
        user = um
        user.username = user_name
        user.password = passwor
        user.email = em
        user.save()
        auth_user = authenticate(username=user_name, password=psw)
        if auth_user is not None:
            auth_login(request, auth_user)
        bb1.Email = em
        bb1.Password = psw
        bb1.First_name = first
        bb1.Last_name = last
        bb1.gst_num = gst_num 
        bb1.user = user
        bb1.Mobile_Number = mn
        bb1.address = addr
        bb1.Image = photo_url  
        bb1.save()
        request.session['logg'] = bb1.id
        gtt = Registration.objects.filter(User_role='admin')
        messages.success(request, 'Profile Updated')
        return render(request, 'login.html', {'gtt': gtt})
    else:
        return render(request, 'admin_home.html')

def del_admin(request, id):
    bb1 = Registration.objects.get(id = id)
    User.objects.get(email = bb1.Email).delete()
    messages.success(request, 'You have successfully resigned from administration')
    return redirect('home')

def admin_rg(request):
    if request.method == 'POST':
        if Registration.objects.filter(User_role='admin').exists():
            messages.success(request, 'An admin account already exists. Registration as admin is not allowed.')
            return redirect('home')
        z = datetime.now().strftime("%Y-%m-%d")
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        mobile_number = request.POST.get('mobile_number')
        email = request.POST.get('email')
        psw = request.POST.get('psw')
        gst_num = request.POST.get('gst_num')
        useragreement = request.POST.get('useragreement')
        address = request.POST.get('address')
        photo = request.FILES.get('photo')
        if not photo:
            messages.error(request, 'Photo upload is required.')
            return render(request, 'register_admin.html')
        fs = FileSystemStorage()
        fs.save(photo.name, photo)
        if Registration.objects.filter(Email=email).exists():
            messages.success(request, 'A user with this email already exists.')
            return render(request, 'register_admin.html')
        user_name = request.POST.get('user_name')
        if User.objects.filter(username=user_name).exists():
            messages.success(request, 'Username is taken. Please try another.')
            return render(request, 'register_admin.html')
        admin_role = 'admin'
        user = User.objects.create_user(username=user_name, email=email, password=psw)
        user.save()
        t = Registration(
            First_name=first_name,
            Last_name=last_name,
            Email=email,
            Password=psw,
            Mobile_Number=mobile_number,
            Registration_date=z,
            gst_num=gst_num,
            Image=photo,
            address=address,
            User_Agreement=useragreement,
            User_role=admin_role, 
            user=user
        )
        t.save()
        messages.success(request, 'You have successfully registered as admin.')
        return redirect('home')
    else:
        return render(request, 'register_admin.html')

def employeess(request):
    bok = Registration.objects.get(id=request.session['logg'])
    all_users = Registration.objects.filter(User_role='employee')
    return render(request, "employees.html", {'all_users': all_users,'bok':bok})


def register_employee(request):
    bok = Registration.objects.get(id=request.session['logg'])
    branch = Branch.objects.all()
    if request.method == 'POST':
        x = datetime.now()
        z = x.strftime("%Y-%m-%d")
        emp_id = request.POST.get('emp_id')
        first_name = request.POST.get('first_name')
        address = request.POST.get('address')
        gst_num = request.POST.get('gst_num')
        mobile_number = request.POST.get('mobile_number')
        email = request.POST.get('email')
        psw = request.POST.get('psw')
        gender = request.POST.get('gender')
        ua = request.POST.get('ua')
        location = request.POST.get('location')
        employee = request.POST.get('employee')
        reg1 = Registration.objects.all()
        for i in reg1:
            if i.Email == email:
                messages.success(request, 'Branch already Registered')
                return render(request, 'register_employee.html')
        user_name = request.POST.get('user_name')
        for t in User.objects.all():
            if t.username == user_name:
                messages.success(request, 'Username taken. Please try another')
                return render(request, 'register_employee.html')
        user = User.objects.create_user(username=user_name, email=email, password=psw)
        user.save()
        t = Registration()
        t.Emp_Id = emp_id
        t.First_name = first_name
        t.address = address
        t.Location = location
        t.Email = email
        t.Password = psw
        t.gst_num=gst_num
        t.Mobile_Number = mobile_number
        t.Registration_date = z
        t.Gender = gender
   
        t.Address = ua
        t.User_role = employee
        t.user = user
        t.save()
        messages.success(request, 'You have Successfully Registered')
        return redirect('employeess')
    else:
        return render(request, 'register_employee.html',{'bok':bok,'branch':branch})

from .models import Branch  # make sure to import your Branch model

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages, auth
from django.contrib.auth.hashers import make_password
from .models import Registration, Branch
from django.contrib.auth.models import User
from django.contrib import messages, auth
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.shortcuts import render, redirect, get_object_or_404
from .models import Registration, Branch
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from .models import Registration, Branch


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from myapp.models import Registration, Branch  # adjust model imports as needed
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from myapp.models import Registration, Branch


def update_employee(request):
    """Update employee profile without expiring session or re-authenticating."""
    
    # ✅ Ensure user session exists
    logg_id = request.session.get('logg')
    if not logg_id:
        messages.warning(request, "Session expired. Please log in again.")
        return redirect('login')  # Redirect safely to login page

    # ✅ Fetch registration and linked user
    reg = get_object_or_404(Registration, id=logg_id)
    user = get_object_or_404(User, email=reg.Email)
    branches = Branch.objects.all()

    if request.method == 'POST':
        # ✅ Get all submitted form fields
        f_name = request.POST.get('first_name')
        designatio = request.POST.get('designatio')
        loc_id = request.POST.get('loc')
        email = request.POST.get('email')
        addres = request.POST.get('addres')
        psw = request.POST.get('psw')
        gst_num = request.POST.get('gst_num')
        user_name = request.POST.get('user_name')
        mobile = request.POST.get('mobile')

        # ✅ Check for duplicate username (excluding current)
        if User.objects.exclude(username=user.username).filter(username=user_name).exists():
            messages.error(request, 'Username already taken. Please try another.')
            return render(request, 'update_employee.html', {
                'bok': reg,
                'um': user,
                'branch': branches
            })

        # ✅ Update User model safely
        user.username = user_name
        user.email = email
        if psw:  # Only update password if provided
            user.password = make_password(psw)
        user.save()

        # ✅ Update Registration model
        reg.First_name = f_name
        reg.Designation = designatio
        reg.loc_id = loc_id
        reg.Email = email
        reg.address = addres
        reg.gst_num = gst_num
        reg.Password = psw if psw else reg.Password  # Keep old password if blank
        reg.Mobile_Number = mobile
        reg.user = user
        reg.save()

        # ✅ Success message, stay on same page
        messages.success(request, 'Profile updated successfully!')
        return redirect('login')

    # ✅ Render page for GET requests
    return render(request, 'update_employee.html', {
        'bok': reg,
        'um': user,
        'branch': branches
    })


def update_employeee(request, employee_id):
    try:
        logged_in_user = get_object_or_404(Registration, id=employee_id)
        user_instance = get_object_or_404(User, email=logged_in_user.Email)

        if request.method == 'POST':
            first_name = request.POST.get('first_name')
            designation = request.POST.get('designation')
            location = request.POST.get('loc')
            email = request.POST.get('email')
            password = request.POST.get('psw')
            photo = request.FILES.get('photo')
            if email and email != user_instance.email:
                if User.objects.filter(email=email).exclude(id=user_instance.id).exists():
                    messages.error(request, 'Email already in use. Please use a different email.')
                    return redirect('update_employeee', employee_id=employee_id)

            user_instance.email = email
            if password:  
                user_instance.password = make_password(password)
            user_instance.save()
            logged_in_user.First_name = first_name
            logged_in_user.Designation = designation
            logged_in_user.Address = location
            logged_in_user.Email = email
            if password:
                logged_in_user.Password = make_password(password) 
            logged_in_user.user = user_instance

            if photo: 
                logged_in_user.Image = photo  
            logged_in_user.save()

            messages.success(request, 'Branch profile updated successfully!')
            return redirect('update_employee')

        return render(request, 'update_employee.html', {'bok': logged_in_user})

    except IntegrityError:
        messages.error(request, 'Database error occurred while updating the profile.')
    except ValidationError as e:
        messages.error(request, f'Validation Error: {e}')
    except Exception as e:
        messages.error(request, f'Error: {e}')

    return redirect('update_employee')


def edit_employee(request, id):
    bok = Registration.objects.get(id=request.session['logg'])
    branch = Branch.objects.all()

    # Get employee record
    employee_data = Registration.objects.get(id=id)

    if request.method == 'POST':

        employee_data.Emp_Id = request.POST.get('emp_id')
        employee_data.First_name = request.POST.get('first_name')
        employee_data.address = request.POST.get('address')
        employee_data.gst_num = request.POST.get('gst_num')
        employee_data.Mobile_Number = request.POST.get('mobile_number')
        employee_data.Email = request.POST.get('email')
        employee_data.Gender = request.POST.get('gender')
        employee_data.Address = request.POST.get('ua')
        employee_data.Location = request.POST.get('location')
        employee_data.User_role = request.POST.get('employee')

        # SMTP Fields
        employee_data.smtp_host = request.POST.get('smtp_host')
        employee_data.smtp_port = request.POST.get('smtp_port')
        employee_data.smtp_email = request.POST.get('smtp_email')
        employee_data.smtp_password = request.POST.get('smtp_password')

        smtp_use_tls = request.POST.get('smtp_use_tls')

        if smtp_use_tls == "True":
            employee_data.smtp_use_tls = True
        else:
            employee_data.smtp_use_tls = False

        # Update password if entered
        psw = request.POST.get('psw')
        if psw:
            employee_data.Password = psw
            employee_data.user.set_password(psw)
            employee_data.user.save()

        # Update Django User table
        user_name = request.POST.get('user_name')

        # Check username already exists for another user
        existing_user = User.objects.filter(username=user_name).exclude(id=employee_data.user.id)

        if existing_user.exists():
            messages.success(request, 'Username taken. Please try another')
            return render(request, 'edit_employee.html', {
                'bok': bok,
                'branch': branch,
                'employee_data': employee_data
            })

        employee_data.user.username = user_name
        employee_data.user.email = request.POST.get('email')
        employee_data.user.save()

        employee_data.save()

        messages.success(request, 'Employee Updated Successfully')
        return redirect('employeess')

    return render(request, 'edit_employee.html', {
        'bok': bok,
        'branch': branch,
        'employee_data': employee_data
    })

        
def del_employee(request, id):
    try:
        employee = get_object_or_404(Registration, id=id)
        user = get_object_or_404(User, email=employee.Email)


        user.delete()
        employee.delete()

        messages.success(request, 'Branch account deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error deleting branch: {e}')
    
    return redirect('employeess')

def adminn_details(request):
    bok = Registration.objects.get(id=request.session['logg'])
    gtt = Registration.objects.filter(User_role = 'admin')
    return render(request, "adminn_details.html",{'bok':bok,'gtt':gtt,})


from datetime import datetime, timedelta
from django.shortcuts import render, redirect


from django.shortcuts import render, redirect
from django.db.models import Sum
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required

from datetime import datetime, timedelta
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from django.db.models import Sum
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import datetime, timedelta
from django.utils.timezone import now

from .models import (
    Registration, Branch, Vendor, Item, Company, Company_line,
    PurchaseOrder, PurchaseOrderLine, Invoice
)
from .decorators import logged_inn4  # Assuming you have this decorator

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import datetime, timedelta
from django.utils.timezone import now
from .models import (
    Registration, Item, Vendor, PurchaseOrder, Invoice,
    Company, Company_line, Branch, PurchaseOrderLine
)
from .decorators import logged_inn4
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import datetime, timedelta
from django.utils.timezone import now
from .models import (
    Registration, Item, Vendor, PurchaseOrder, Invoice,
    Company, Company_line, Branch, PurchaseOrderLine
)
from .decorators import logged_inn4

from datetime import datetime, timedelta
from django.db.models import Sum
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from .models import Registration, Item, Vendor, PurchaseOrder, PurchaseOrderLine, Invoice, Company, Company_line, Branch


@login_required
@logged_inn4
def employee_home(request):
    user_id = request.session.get('logg')
    if not user_id:
        return redirect('login')

    bok = get_object_or_404(Registration, id=user_id)
    if bok.payment_expiry and bok.payment_expiry < now():
        bok.payment_status = False
        bok.save()

    # ------------------- Date Filter Logic -------------------
    date_filter = request.GET.get('filter', 'monthly')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    today = now().date()
    start_date, end_date = None, None

    if date_filter == 'daily':
        start_date = end_date = today
    elif date_filter == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif date_filter == 'monthly':
        start_date = today.replace(day=1)
        end_date = today
    elif date_filter == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif date_filter == 'custom' and start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = end_date = None

    # ------------------- Base filters -------------------
    filters = {
        'user_id': user_id
    }

    def apply_date_filter(qs, date_field):
        if start_date and end_date:
            return qs.filter(**{f"{date_field}__range": (start_date, end_date)})
        return qs

    # ------------------- Fetch data with filters -------------------
    items = apply_date_filter(Item.objects.filter(**filters), 'created_at').select_related('vendor')
    vendors = apply_date_filter(Vendor.objects.filter(**filters), 'created_at')
    po = apply_date_filter(PurchaseOrder.objects.filter(**filters), 'dateField')
    invoices = apply_date_filter(Invoice.objects.filter(**filters), 'dateField')  # <-- filter by dateField now
    company = Company.objects.filter(**filters)
    company_lines = apply_date_filter(Company_line.objects.filter(**filters), 'created_at')
    bran = Branch.objects.all()
    gtt = Registration.objects.filter(User_role='employee', id=user_id)
    ven = vendors

    # ------------------- Totals -------------------
    total_registrations = Registration.objects.filter(id=user_id).count()
    total_vendors = vendors.count()
    total_items = items.count()
    total_branches = bran.count()
    total_companies = company.count()
    total_company_lines = company_lines.count()
    total_purchase_orders = po.count()

    # ------------------- Orders Today -------------------
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    total_purchase_order_items_today = PurchaseOrderLine.objects.filter(
        purchase_order__user_id=user_id,
        purchase_order__dateField__gte=today_start,
        purchase_order__dateField__lt=today_end
    ).count()

    # ------------------- Total Revenue -------------------
    po_lines = PurchaseOrderLine.objects.filter(purchase_order__user_id=user_id)
    if start_date and end_date:
        po_lines = po_lines.filter(purchase_order__dateField__range=(start_date, end_date))
    total_purchase_order_amount = po_lines.aggregate(total=Sum('total_amount'))['total'] or 0

    # ------------------- Invoice Aggregation -------------------
    invoice_totals = invoices.aggregate(
        total_invoice_subtotal=Sum('sub_total'),
        total_invoice_sgst=Sum('total_sgst'),
        total_invoice_cgst=Sum('total_cgst'),
        total_invoice_amount=Sum('total_amount'),
    )

    # ------------------- Context -------------------
    context = {
        'bok': bok,
        'k': gtt,
        'gtt': gtt,
        'po': po,
        'items': items,
        'vendors': vendors,
        'invoices': invoices,
        'ven': ven,
        'bran': bran,
        'total_registrations': total_registrations,
        'total_vendors': total_vendors,
        'total_items': total_items,
        'total_branches': total_branches,
        'total_companies': total_companies,
        'total_company_lines': total_company_lines,
        'total_purchase_order_items': total_purchase_orders,
        'total_purchase_order_items_today': total_purchase_order_items_today,
        'total_purchase_order_amount': total_purchase_order_amount,
        'invoice_totals': invoice_totals,
        'filter': date_filter,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }

    return render(request, 'employee_home.html', context)




def view_staffs(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    view_s =  Registration.objects.filter(User_role = 'employee')
    return render(request, 'view_staffs.html',{'bok':bok,'view_s':view_s} )



def home(request):
    return render(request,'home.html',)

from django.shortcuts import redirect
from django.contrib import auth
from django.utils import timezone
from django.shortcuts import redirect
from django.contrib import auth
from django.utils import timezone
from django.shortcuts import redirect
from django.contrib import auth
from datetime import datetime

def logout(request):
    user = request.user
    if user.is_authenticated:
        # Save logout time as string in the User's last_name field (system/local time)
        logout_time = datetime.now().strftime("%d-%m-%Y %I:%M %p")
        user.last_name = logout_time
        user.save()

        # Clear custom session key if exists
        request.session.pop('logg', None)

        # Logout the user
        auth.logout(request)

    return redirect('login')



def branch_create(request):
    if request.method == 'POST':
    
        location = request.POST.get('location')
        Branch.objects.create( location=location)
        return redirect('branch_list')
    return render(request, 'branch_form.html', {'action': 'Create'})

def branch_list(request):
    branches = Branch.objects.all()
    return render(request, 'branch_list.html', {'branches': branches})

def branch_detail(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    return render(request, 'branch_detail.html', {'branch': branch})


def branch_update(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':

        branch.location = request.POST.get('location')
        branch.save()
        return redirect('branch_list')
    return render(request, 'update_branch.html', {'branch': branch, 'action': 'Update'})

def branch_delete(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':
        branch.delete()
        return redirect('branch_list')
    return render(request, 'branch_confirm_delete.html', {'branch': branch})


from django.contrib.auth.decorators import login_required

@login_required
def company_list(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    companies = Company.objects.filter(user=request.user)
    items = Item.objects.filter(user=request.user)
    vendors = Vendor.objects.filter(user=request.user)
    branches = Branch.objects.all()

    return render(request, 'company_list.html', {
        'companies': companies,
        'items': items,
        'vendors': vendors,
        'branches': branches,
        'bok':bok,
    })

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Item




import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import get_object_or_404

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
import json



from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse
from django.shortcuts import get_object_or_404

import json


import json
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db import transaction

from django.db import transaction
from .models import PurchaseOrder, PurchaseOrderLine, Vendor, Branch, Item
from django.contrib.auth.models import User
from decimal import Decimal

from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from .models import Vendor, Branch, Item, PurchaseOrder, PurchaseOrderLine


from django.views.decorators.csrf import csrf_exempt  # Only in development!
from django.shortcuts import render, redirect
from django.contrib import messages
from decimal import Decimal
from .models import Vendor, Branch, Item, PurchaseOrder, PurchaseOrderLine

from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.contrib import messages
from decimal import Decimal
from .models import PurchaseOrder, PurchaseOrderLine, Vendor, Branch, Item


from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.contrib import messages
from decimal import Decimal
from .models import Vendor, Branch, Item, PurchaseOrder, PurchaseOrderLine
from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Vendor, Branch, Item, PurchaseOrder, PurchaseOrderLine


from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Vendor, Item, PurchaseOrder, PurchaseOrderLine
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.contrib import messages
from .models import PurchaseOrder, PurchaseOrderLine, Vendor, Item
from django.contrib.auth.models import User

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from decimal import Decimal
from .models import Vendor, Item, PurchaseOrder, PurchaseOrderLine

from decimal import Decimal
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from .models import PurchaseOrder, PurchaseOrderLine, Vendor, Item
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from decimal import Decimal
import json
from .models import Vendor, Item, PurchaseOrder, PurchaseOrderLine
from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import Vendor, PurchaseOrder, PurchaseOrderLine, Item

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal, InvalidOperation
from .models import PurchaseOrder, PurchaseOrderLine, Vendor, Item


from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse
from .models import ConsumedPurchaseOrder, ConsumedPurchaseOrderLine, Vendor, Item

from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.shortcuts import redirect
from decimal import Decimal, InvalidOperation
from .models import Vendor, Item, ConsumedPurchaseOrder, ConsumedPurchaseOrderLine


from django.shortcuts import redirect
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from django.views.decorators.csrf import csrf_exempt

from django.shortcuts import redirect
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from django.views.decorators.csrf import csrf_exempt

from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from .models import Vendor, Item, ConsumedPurchaseOrder, ConsumedPurchaseOrderLine

from decimal import Decimal, InvalidOperation
from django.shortcuts import redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
from django.shortcuts import redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from .models import Vendor, Item, ConsumedPurchaseOrder, ConsumedPurchaseOrderLine


@csrf_exempt
def create_consumed_order(request):
    if request.method != 'POST':
        messages.warning(request, "Invalid request method.")
        return redirect('view_consumed_orders')

    try:
        # --- Vendor Validation ---
        vendor_id = request.POST.get('vendor_id')
        if not vendor_id or not vendor_id.isdigit():
            messages.error(request, "Invalid vendor ID.")
            return redirect('view_consumed_orders')

        try:
            vendor = Vendor.objects.get(id=int(vendor_id))
        except Vendor.DoesNotExist:
            messages.error(request, "Vendor not found.")
            return redirect('view_consumed_orders')

        user = request.user if request.user.is_authenticated else None

        # --- Fetch line item fields ---
        item_ids = request.POST.getlist('item_id[]')
        prices = request.POST.getlist('price[]')
        packs = request.POST.getlist('pack[]')
        stocks = request.POST.getlist('stock[]')
        qtys = request.POST.getlist('qty[]')
        hsn_codes = request.POST.getlist('hsn[]')
        gst_vals = request.POST.getlist('gst[]')

        line_count = len(item_ids)
        all_lists = [prices, packs, stocks, qtys, hsn_codes, gst_vals]
        if not all(len(lst) == line_count for lst in all_lists):
            messages.error(request, "Mismatch in line item data.")
            return redirect('view_consumed_orders')

        # --- Validate stock availability ---
        for i in range(line_count):
            try:
                item = Item.objects.get(id=item_ids[i])
                qty = Decimal(qtys[i])
                item.stock_balance = item.stock_balance or Decimal('0.00')
                if qty > item.stock_balance:
                    messages.warning(
                        request,
                        f"Not enough stock for item '{item.name}'. "
                        f"Available: {item.stock_balance}, Requested: {qty}"
                    )
                    return redirect('view_consumed_orders')
            except (Item.DoesNotExist, InvalidOperation, ValueError):
                messages.error(request, f"Invalid item or quantity at line {i + 1}.")
                return redirect('view_consumed_orders')

        # --- Handle manual date field ---
        created_at_str = request.POST.get('created_at', '').strip()
        if created_at_str:
            try:
                dateField = datetime.strptime(created_at_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
                return redirect('view_consumed_orders')
        else:
            dateField = date.today()  # Default to today if not provided

        # --- Create Consumed Purchase Order ---
        company_name = request.POST.get('company', '').strip()
        invoice_number = request.POST.get('invoice_number', '').strip()

        consumed_po = ConsumedPurchaseOrder.objects.create(
            vendor=vendor,
            company=company_name,
            dateField=dateField,  # ✅ New date field
            invoice_number=invoice_number,
            user=user,
            created_at=dateField
        )

        # --- Create line items and update stock ---
        for i in range(line_count):
            item = Item.objects.get(id=item_ids[i])
            price = Decimal(prices[i])
            qty = Decimal(qtys[i])
            gst = Decimal(gst_vals[i])
            pack = packs[i]
            stock = int(float(stocks[i]))
            hsn = hsn_codes[i]

            amount = price * qty
            total_amount = amount + (amount * gst / 100)

            ConsumedPurchaseOrderLine.objects.create(
                consumed_order=consumed_po,
                item=item,
                price=price,
                pack=pack,
                stock=stock,
                qty=qty,
                gst=gst,
                hsn=hsn,
                amount=amount,
                total_amount=total_amount,
            )

            # Update stock balance
            item.stock_balance -= qty
            item.save()
        log_usage(
            request.user,
            'CONSUMED_ORDER_CREATE'
        )
        messages.success(request, "Consumed created successfully.")
        return redirect('custom_consumed')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('view_consumed_orders')

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import PurchaseOrderLine


from django.http import JsonResponse
from .models import PurchaseOrderLine
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import PurchaseOrderLine
from decimal import Decimal
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import Vendor, PurchaseOrderLine, Item

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import PurchaseOrderLine, Vendor
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import Item, Vendor
from django.http import JsonResponse
from .models import Item, Vendor

# Get details of a single item (or multiple if needed)
from django.http import JsonResponse
from .models import Item, Vendor
from django.http import JsonResponse
from .models import Item, Vendor

from django.http import JsonResponse
from .models import Item, Vendor
from django.contrib.auth.models import User  # or your custom user model
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Item


def get_item_details(request):
    """
    Fetch details of a single item by item_id.
    Returns JSON with all relevant item fields.
    """
    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'Item ID is required'}, status=400)

    try:
        item = get_object_or_404(Item, id=item_id)

        return JsonResponse({
            'item_id': item.id,
            'name': item.name or '',
            'company': item.company_name or (item.vendor.name if item.vendor else ''),
            'company_name': item.company_name or (item.vendor.name if item.vendor else ''),
            'price': str(item.item_price) if item.item_price is not None else '',
            'item_price': str(item.item_price) if item.item_price is not None else '',
            'pack': item.pack_size or '',
            'pack_size': item.pack_size or '',
            'stock': str(item.stock_balance) if item.stock_balance is not None else '',
            'stock_balance': str(item.stock_balance) if item.stock_balance is not None else '',
            'gst': str(item.gst) if item.gst is not None else '',
            'hsn': item.hsn or '',
            'vendor_id': item.vendor.id if item.vendor else None,
            'vendor': {
                'id': item.vendor.id,
                'name': item.vendor.name
            } if item.vendor else None
        })

    except Exception as e:
        return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)



def get_vendor_details(request):
    """
    Fetch all items for a vendor by vendor_id.
    Returns JSON list of items.
    """
    vendor_id = request.GET.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'error': 'Vendor ID is required'}, status=400)

    try:
        vendor = Vendor.objects.get(id=vendor_id)
        items_data = []

        for item in vendor.items.all():  # Uses related_name='items'
            items_data.append({
                'item_id': item.id,
                'name': item.name or '',
                'company': item.company_name or vendor.name,
                'price': str(item.item_price) if item.item_price is not None else '',
                'pack': item.pack_size or '',
                'stock': str(item.stock_balance) if item.stock_balance is not None else '',
                'gst': item.gst or '',
                'hsn': item.hsn or ''
            })

        return JsonResponse({'items': items_data})

    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)





def get_companies(request):
    type_ = request.GET.get('type')
    id_ = request.GET.get('id')

    if type_ == 'item':
        try:
            item = Item.objects.get(id=id_)
            items = Item.objects.filter(name=item.name)
        except Item.DoesNotExist:
            return JsonResponse([], safe=False)
    elif type_ == 'vendor':
        items = Item.objects.filter(vendor_id=id_)
    else:
        return JsonResponse([], safe=False)

    data = [
        {
            'company': item.company_name,
            'name':item.name,
            'price': str(item.item_price),
            'pack': item.pack_size,
            'stock': item.stock_balance,
            'gst': item.gst,
            'hsn': item.hsn,
            'item_id': item.id,
            'vendor_id': item.vendor_id,  # Include vendor ID
            'vendor_name': item.vendor.name  # Include vendor name
        }
        for item in items
    ]

    return JsonResponse(data, safe=False)
from django.views.decorators.http import require_GET

@require_GET
def get_vendor_data(request, vendor_id):
    """
    Fetch vendor related data
    """
    try:
        vendor = Vendor.objects.get(pk=vendor_id)

        return JsonResponse({
            "vendor_id": vendor.id,
            "company_name": vendor.company_name,
        })

    except Vendor.DoesNotExist:
        return JsonResponse(
            {"error": "Vendor not found"},
            status=404
        )


@require_GET
def get_item_data(request, item_id):
    """
    Fetch item related data
    """
    try:
        item = Item.objects.get(pk=item_id)

        return JsonResponse({
            "item_id": item.id,
            "name": item.name,          # ✅ Item name added
            "price": item.price,
            "pack": item.pack,
            "stock": item.stock,
            "gst": item.gst,
            "hsn": item.hsn,
        })

    except Item.DoesNotExist:
        return JsonResponse(
            {"error": "Item not found"},
            status=404
        )


def get_items_by_vendor(request, vendor_id):
    items = Item.objects.filter(vendor_id=vendor_id)
    item_list = [
        {
            'item_id': item.id,
            'name': item.name,
            'company': item.company_name,
            'price': str(item.item_price),
            'pack': item.pack_size,
            'gst': item.gst,
            'hsn': item.hsn,
            'stock': item.stock_balance,
            'vendor_id': item.vendor_id,
        }
        for item in items
    ]
    return JsonResponse(item_list, safe=False)
@require_GET
def get_vendor_items(request):
    vendor_id = request.GET.get('vendor_id')

    if not vendor_id:
        return JsonResponse({'error': 'Vendor ID missing'}, status=400)

    items = Item.objects.filter(vendor_id=vendor_id)

    return JsonResponse({
        'items': [
            {
                'id': item.id,
                'name': item.name
            }
            for item in items
        ]
    })

from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from datetime import datetime, date
from .models import Vendor, Item, ConsumedPurchaseOrder, ConsumedPurchaseOrderLine
from decimal import Decimal
from datetime import datetime, date
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt

from .models import Registration, ConsumedPurchaseOrder, ConsumedPurchaseOrderLine, Item, Vendor

@csrf_exempt
def update_consumed_order(request, order_id):
    # Get logged-in user
    bok = get_object_or_404(Registration, id=request.session.get('logg'))

    # Get order
    po = get_object_or_404(ConsumedPurchaseOrder, id=order_id)

    if request.method == 'GET':
        vendors = Vendor.objects.all()
        items = Item.objects.all()
        lines = ConsumedPurchaseOrderLine.objects.filter(consumed_order=po)

        return render(request, 'update_consumed_order.html', {
            'order': po,
            'vendors': vendors,
            'items': items,
            'lines': lines,
            'current_date': po.dateField or date.today(),
            'bok': bok,
        })

    elif request.method == 'POST':
        try:
            # -----------------------
            # Update basic order info
            # -----------------------
            vendor_id = request.POST.get('vendor_id')
            po.company = request.POST.get('company', '').strip()
            po.invoice_number = request.POST.get('invoice_number', '').strip()

            if vendor_id and vendor_id.isdigit():
                po.vendor = get_object_or_404(Vendor, id=int(vendor_id))
            else:
                po.vendor = None  # allow no vendor

            # Update date
            created_at_str = request.POST.get('created_at', '').strip()
            if created_at_str:
                try:
                    po.dateField = datetime.strptime(created_at_str, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
                    return redirect('update_consumed_order', order_id=order_id)
            else:
                po.dateField = date.today()

            po.save()

            # -----------------------
            # Restore old stock & delete old lines
            # -----------------------
            old_lines = ConsumedPurchaseOrderLine.objects.filter(consumed_order=po)
            for line in old_lines:
                if line.item:
                    line.item.stock_balance += line.qty
                    line.item.save()
            old_lines.delete()

            # -----------------------
            # Process new lines
            # -----------------------
            item_ids = request.POST.getlist('item_id[]')
            names = request.POST.getlist('name[]')
            companies = request.POST.getlist('company[]')
            prices = request.POST.getlist('price[]')
            packs = request.POST.getlist('pack[]')
            stocks = request.POST.getlist('stock[]')
            qtys = request.POST.getlist('qty[]')
            hsn_codes = request.POST.getlist('hsn[]')
            gst_vals = request.POST.getlist('gst[]')

            for i in range(len(item_ids)):
                item_id = item_ids[i]
                # Check if manual item
                if item_id.startswith('manual_'):
                    # Create new Item
                    item = Item.objects.create(
                        name=names[i],
                        company_name=companies[i],
                        item_price=Decimal(prices[i]),
                        pack_size=packs[i],
                        stock_balance=int(float(stocks[i])),
                        gst=Decimal(gst_vals[i]),
                        hsn=hsn_codes[i],
                    )
                else:
                    item = get_object_or_404(Item, id=item_id)

                price = Decimal(prices[i])
                qty = Decimal(qtys[i])
                gst = Decimal(gst_vals[i])
                pack = packs[i]
                stock = int(float(stocks[i]))
                hsn = hsn_codes[i]

                amount = price * qty
                total_amount = amount + (amount * gst / 100)

                # Create line
                ConsumedPurchaseOrderLine.objects.create(
                    consumed_order=po,
                    item=item,
                    price=price,
                    pack=pack,
                    stock=stock,
                    qty=qty,
                    gst=gst,
                    hsn=hsn,
                    amount=amount,
                    total_amount=total_amount,
                )

                # Deduct stock
                item.stock_balance -= qty
                item.save()

            messages.success(request, "Consumed Order updated successfully.")
            return redirect('custom_consumed')

        except Exception as e:
            messages.error(request, f"Error updating order: {str(e)}")
            return redirect('update_consumed_order', order_id=order_id)

    return HttpResponse("Invalid request method.", status=405)

from django.http import JsonResponse
from .models import Vendor, Item
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import Vendor, Item




from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from decimal import Decimal
from django.http import HttpResponse
from .models import ConsumedPurchaseOrder, ConsumedPurchaseOrderLine

@csrf_exempt
def delete_consumed_order(request, consumed_po_id):
    if request.method == 'POST':
        try:
            po = get_object_or_404(ConsumedPurchaseOrder, id=consumed_po_id)

            # Restore stock for each line item
            lines = ConsumedPurchaseOrderLine.objects.filter(consumed_order=po)
            for line in lines:
                item = line.item
                if item:
                    restored_qty = line.qty or Decimal('0.00')
                    item.stock_balance = (item.stock_balance or Decimal('0.00')) + restored_qty
                    item.save()

            # Delete lines and then the main order
            lines.delete()
            po.delete()

            messages.success(request, "Deleted successfully.")
            return redirect('custom_consumed')

        except Exception as e:
            return HttpResponse(f"Error deleting Consumed Order: {str(e)}", status=500)

    return HttpResponse("Invalid request method.", status=405)



from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import redirect
from django.urls import reverse
from .models import Vendor, PurchaseOrder, PurchaseOrderLine, Item
from django.db.models import Max
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.shortcuts import redirect
from .models import Vendor, PurchaseOrder, Item, PurchaseOrderLine


@csrf_exempt
def edit_purchase_order(request, po_id):
    po = get_object_or_404(PurchaseOrder, id=po_id)
    if request.method == 'POST':
        data = request.POST
        try:
            vendor = Vendor.objects.get(id=data.get('vendor_id'))

            # Update PO fields
            po.vendor = vendor
            po.invoice_number = data.get('invoice_number', '')
            po.company = data.get('company', '')
            po.save()

            # Delete existing lines
            po.order_lines.all().delete()

            # Re-create lines
            item_ids = data.getlist('item_id[]')
            prices = data.getlist('price[]')
            packs = data.getlist('pack[]')
            stocks = data.getlist('stock[]')
            qtys = data.getlist('qty[]')
            gsts = data.getlist('gst[]')
            hsns = data.getlist('hsn[]')

            for i in range(len(item_ids)):
                item = Item.objects.get(id=item_ids[i])
                price = Decimal(prices[i]) if prices[i] else Decimal('0')
                qty = Decimal(qtys[i]) if qtys[i] else Decimal('0')
                gst = Decimal(gsts[i]) if gsts[i] else Decimal('0')
                amount = price * qty
                total_amount = amount + (amount * gst / 100)

                PurchaseOrderLine.objects.create(
                    purchase_order=po,
                    item=item,
                    price=price,
                    pack=packs[i],
                    stock=int(stocks[i]) if stocks[i] else 0,
                    qty=qty,
                    gst=gst,
                    hsn=hsns[i],
                    amount=amount,
                    total_amount=total_amount,
                )

            messages.success(request, "Purchase Order updated successfully.")
            return redirect('view_purchase_orders')

        except Exception as e:
            return HttpResponse(f"Error updating Purchase Order: {str(e)}")

    vendors = Vendor.objects.all()
    items = Item.objects.all()
    return render(request, 'edit_purchase_order.html', {
        'po': po,
        'vendors': vendors,
        'items': items,
    })


from django.http import JsonResponse

@csrf_exempt
def delete_purchase_order(request, po_id):
    if request.method == 'POST':
        try:
            po = get_object_or_404(PurchaseOrder, id=po_id)
            po.delete()
            messages.success(request, "Purchase Order deleted successfully.")
            return redirect('view_purchase_orders')
        except Exception as e:
            return HttpResponse(f"Error deleting Purchase Order: {str(e)}")




@csrf_exempt
def edit_consumed_order(request, consumed_po_id):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    consumed_po = get_object_or_404(ConsumedPurchaseOrder, id=consumed_po_id)

    if request.method == 'POST':
        data = request.POST
        try:
            vendor = Vendor.objects.get(id=data.get('vendor_id'))

            # Update main PO fields
            consumed_po.vendor = vendor
            consumed_po.invoice_number = data.get('invoice_number', '')
            consumed_po.company = data.get('company', '')
            consumed_po.save()

            # Delete existing lines
            consumed_po.consumed_lines.all().delete()

            # Re-create lines
            item_ids = data.getlist('item_id[]')
            prices = data.getlist('price[]')
            packs = data.getlist('pack[]')
            stocks = data.getlist('stock[]')
            qtys = data.getlist('qty[]')
            gsts = data.getlist('gst[]')
            hsns = data.getlist('hsn[]')

            for i in range(len(item_ids)):
                item = Item.objects.get(id=item_ids[i])
                price = Decimal(prices[i]) if prices[i] else Decimal('0')
                qty = Decimal(qtys[i]) if qtys[i] else Decimal('0')
                gst = Decimal(gsts[i]) if gsts[i] else Decimal('0')
                amount = price * qty
                total_amount = amount + (amount * gst / 100)

                ConsumedPurchaseOrderLine.objects.create(
                    consumed_order=consumed_po,
                    item=item,
                    price=price,
                    pack=packs[i],
                    stock=int(stocks[i]) if stocks[i] else 0,
                    qty=qty,
                    gst=gst,
                    hsn=hsns[i],
                    amount=amount,
                    total_amount=total_amount,
                )

            messages.success(request, "Consumed Order updated successfully.")
            return redirect('view_consumed_orders')  # Update this with your actual URL name

        except Exception as e:
            return HttpResponse(f"Error updating Consumed Order: {str(e)}")

    vendors = Vendor.objects.all()
    items = Item.objects.all()
    return render(request, 'edit_consumed_order.html', {
        'po': consumed_po,
        'vendors': vendors,
        'items': items,
        'bok':bok,
    })


        
def view_consumed_orders(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    companies = Company.objects.filter(user=request.user)
    items = Item.objects.filter(user=request.user)
    vendors = Vendor.objects.filter(user=request.user)
    branches = Branch.objects.all()
    consumed_orders = ConsumedPurchaseOrder.objects.all().order_by('-created_at')
    return render(request, 'view_consumed_orders.html', {
        'consumed_orders': consumed_orders,
           'companies': companies,
        'items': items,
        'vendors': vendors,
        'branches': branches,
        'bok':bok,
    })

from datetime import date, timedelta
from django.shortcuts import render
from .models import PurchaseOrder, Vendor
from django.shortcuts import render, get_object_or_404
from datetime import date, timedelta
from .models import Registration, PurchaseOrder, Vendor

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from .models import Registration, PurchaseOrder, Vendor
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from datetime import date, timedelta

from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from .models import PurchaseOrder, Vendor, Registration

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import date, timedelta
from .models import PurchaseOrder, Vendor, Registration

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import date, timedelta
from .models import PurchaseOrder, Vendor, Registration

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from .models import PurchaseOrder, Vendor

from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import PurchaseOrder, Vendor, Registration

from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import PurchaseOrder, Vendor, Registration
from datetime import date, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils.http import urlencode
from django.db import transaction




from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import PurchaseOrder

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import PurchaseOrder

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import PurchaseOrder

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import PurchaseOrder
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from num2words import num2words
from .models import PurchaseOrder, Registration




from decimal import Decimal
from num2words import num2words



from django.views.decorators.http import require_POST
from django.contrib import messages


# views.py
from django.shortcuts import redirect
from django.contrib import messages
from .models import PurchaseOrder
from datetime import date, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from .models import PurchaseOrder, Vendor, Registration
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from datetime import date, timedelta
from .models import PurchaseOrder, Vendor, Registration


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from datetime import date, timedelta
import json
from .models import PurchaseOrder, Vendor, Registration

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from datetime import date, timedelta
import json
from .models import PurchaseOrder, Vendor, Registration

@login_required
def view_purchase_orders(request):
    bok = get_object_or_404(Registration, user=request.user)

    # ---------------- FILTER PARAMETERS ----------------
    vendor_name = request.GET.get('vendor', '').strip()
    status = request.GET.get('status', '').strip()
    date_range = request.GET.get('date_range', 'today').strip()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except:
        per_page = 10

    # ---------------- BASE QUERY ----------------
    queryset = PurchaseOrder.objects.filter(user=bok.user).order_by('-dateField').distinct()

    # ---------------- FILTERS ----------------
    if vendor_name:
        queryset = queryset.filter(order_lines__item__vendor__name__iexact=vendor_name)

    if status == 'place order':
        queryset = queryset.filter(invoice__isnull=True)
    elif status == 'in progress':
        queryset = queryset.filter(invoice__status__in=['Pending', 'Unpaid'])
    elif status == 'completed':
        queryset = queryset.filter(invoice__status='Paid')

    today = date.today()
    if date_range == 'today':
        queryset = queryset.filter(dateField=today)
    elif date_range == 'week':
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        queryset = queryset.filter(dateField__range=[start_week, end_week])
    elif date_range == 'month':
        queryset = queryset.filter(dateField__year=today.year, dateField__month=today.month)
    elif date_range == 'year':
        queryset = queryset.filter(dateField__year=today.year)
    elif date_range == 'custom' and date_from and date_to:
        queryset = queryset.filter(dateField__range=[date_from, date_to])

    # ---------------- PAGINATION ----------------
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    vendors = Vendor.objects.all()

    # Keep all current filters (including page) for links and bulk delete
    current_filters = request.GET.copy()
    if 'page' in current_filters:
        current_filters.pop('page')

    context = {
        'purchase_orders': page_obj,
        'page_obj': page_obj,
        'vendors': vendors,
        'bok': bok,
        'vendor_name': vendor_name,
        'status': status,
        'date_range': date_range,
        'date_from': date_from,
        'date_to': date_to,
        'per_page': per_page,
        'per_page_options': [10, 25, 50, 100],
        'current_filters': current_filters.urlencode(),  # filters for pagination links
    }

    return render(request, 'display_purchase_orders.html', context)


@login_required
def bulk_delete_purchase_orders(request):
    if request.method == "POST":
        try:
            # Handle JSON body from fetch
            body_unicode = request.body.decode('utf-8')
            body_data = json.loads(body_unicode)
            selected_ids = body_data.get('ids', [])
        except:
            # Fallback to traditional POST form
            selected_ids = request.POST.getlist('selected_orders')

        if selected_ids:
            deleted_count = PurchaseOrder.objects.filter(id__in=selected_ids).delete()[0]
            messages.success(request, f"{deleted_count} purchase orders deleted successfully.")

    # Redirect back to the same page with current filters
    referer = request.META.get('HTTP_REFERER', '')
    if referer:
        return redirect(referer)
    return redirect('view_purchase_orders')

from django.core.mail import EmailMessage
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.conf import settings

def send_purchase_order_email(request):

    if request.method == 'POST':

        try:

            vendor_id = request.POST.get('vendor_id')
            pdf_file = request.FILES.get('pdf_file')

            vendor = get_object_or_404(Vendor, id=vendor_id)

            registration = get_object_or_404(
                Registration,
                user=request.user
            )

            branch_name = registration.First_name or "Company"

            # DYNAMIC SMTP SETTINGS

            settings.EMAIL_HOST = registration.smtp_host or 'smtp.gmail.com'

            settings.EMAIL_PORT = 587

            settings.EMAIL_USE_TLS = True

            settings.EMAIL_HOST_USER = registration.smtp_email

            settings.EMAIL_HOST_PASSWORD = registration.smtp_password

            # EMAIL

            email = EmailMessage(
                subject="Purchase Order Invoice",

                body=f"""
Hello {vendor.name},

Please find attached the purchase order invoice.

Regards,
{branch_name}
                """,

                from_email=settings.EMAIL_HOST_USER,

                to=[vendor.email],
            )

            # ATTACH PDF

            if pdf_file:
                email.attach(
                    "purchase_invoice.pdf",
                    pdf_file.read(),
                    "application/pdf"
                )

            # SEND EMAIL

            email.send()

            return JsonResponse({
                "status": "success"
            })

        except Exception as e:

            print("EMAIL ERROR:", str(e))

            return JsonResponse({
                "status": "failed",
                "message": str(e)
            }, status=500)

    return JsonResponse({
        "status": "failed"
    }, status=400)


# views.py

import os
from django.http import JsonResponse
from django.core.mail import EmailMessage
from django.shortcuts import get_object_or_404
from django.core.files.storage import default_storage
from django.conf import settings

from .models import Invoice  # Adjust import to your app structure


def send_invoice_email(request, invoice_id):
    """
    Send invoice PDF to vendor's email.
    """
    if request.method == 'POST':
        pdf_file = request.FILES.get('pdf_file')
        invoice = get_object_or_404(Invoice, id=invoice_id)
        vendor = invoice.vendor

        if not vendor.email:
            return JsonResponse({'error': 'Vendor email not found.'}, status=400)

        if pdf_file:
            subject = f"Invoice #{invoice.invoice_number}"
            body = "Please find attached your invoice."
            email = EmailMessage(subject, body, 'you@example.com', [vendor.email])
            email.attach(f"invoice_{invoice.invoice_number}.pdf", pdf_file.read(), 'application/pdf')
            email.send()
            log_usage(
                request.user,
                'EMAIL_SENT'
            )

            return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'failed'}, status=400)


def upload_invoice_for_whatsapp(request, invoice_id):
    """
    Accept invoice PDF blob and return a public file URL for WhatsApp sharing.
    """
    if request.method == 'POST':
        pdf_file = request.FILES.get('pdf_file')
        invoice = get_object_or_404(Invoice, id=invoice_id)

        if pdf_file:
            filename = f"invoice_{invoice.invoice_number}.pdf"
            file_path = os.path.join("invoices", filename)  # saves to MEDIA_ROOT/invoices/
            saved_path = default_storage.save(file_path, pdf_file)

            file_url = request.build_absolute_uri(settings.MEDIA_URL + saved_path)
            return JsonResponse({'status': 'success', 'file_url': file_url})

    return JsonResponse({'status': 'failed'}, status=400)

from django.core.files.storage import default_storage
from django.http import JsonResponse
import uuid

def upload_invoice_for_whatsapp(request):
    if request.method == 'POST' and request.FILES.get('pdf_file'):
        pdf_file = request.FILES['pdf_file']
        filename = f"invoices/{uuid.uuid4()}.pdf"
        path = default_storage.save(filename, pdf_file)
        file_url = request.build_absolute_uri(default_storage.url(path))
        log_usage(
            request.user,
            'WHATSAPP_SHARE'
        )

        return JsonResponse({'success': True, 'file_url': file_url})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

from django.shortcuts import get_object_or_404, redirect

from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse

from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from decimal import Decimal

from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from decimal import Decimal


def purchase_order_item_delete(request, pk):
    try:
        po_item = get_object_or_404(PurchaseOrderLine, pk=pk)

        if request.method == 'POST':
            item = po_item.item


            # Delete the purchase order item
            po_item.delete()

        return redirect('view_po')  # Redirect regardless of GET or POST

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error: {str(e)}'}, status=500)





def purchase_order_invoice(request, pk):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    po = get_object_or_404(PurchaseOrder, pk=pk)
    order_lines = po.order_lines.all()

    # Calculate total including GST
    total_incl_gst = Decimal(0)
    for line in order_lines:
        rate = Decimal(line.amount)
        qty = Decimal(line.qty)
        gst = Decimal(line.gst)
        gross = rate * qty
        gst_amt = gross * gst / Decimal(100)
        total_incl_gst += gross + gst_amt

    amount_in_words = num2words(
        total_incl_gst.quantize(Decimal('0.01')),
        to='currency',
        lang='en_IN'
    ).replace('euro', 'Rupees').replace('cents', 'Only')

    return render(request, 'purchase_order_invoice.html', {
        'purchase_order': po,
        'order_lines': order_lines,
        'bok': bok,
        'amount_in_words': amount_in_words,
        
    })



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal
from num2words import num2words

from .models import PurchaseOrder, PurchaseOrderApproval, Invoice, InvoiceLineItem
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST
from decimal import Decimal
from num2words import num2words

from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal
from num2words import num2words
from django.views.decorators.http import require_POST

from .models import PurchaseOrder, PurchaseOrderApproval, Invoice, InvoiceLineItem


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST
from decimal import Decimal
from num2words import num2words

from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval


@require_POST
def approve_po_items(request, pk):
    """
    Approve Purchase Order items, update approval status, and generate invoice.
    Prevents duplicate invoice creation.
    """
    # Step 1: Get the Purchase Order
    po = get_object_or_404(PurchaseOrder, pk=pk)

    # Step 2: Prevent duplicate invoice creation
    existing_invoice = Invoice.objects.filter(purchase_order=po).first()
    if existing_invoice:
        messages.warning(request, "Invoice already exists for this Purchase Order.")
        return redirect('purchase_order_invoice', pk=existing_invoice.pk)

    # Step 3: Create or update approval status for each order line
    for item in po.order_lines.all():
        approval, created = PurchaseOrderApproval.objects.get_or_create(
            purchase_order_line=item,
            defaults={
                'approval_status': 'Waiting for Approval',
                'user': request.user
            }
        )
        if not created:
            approval.approval_status = 'Waiting for Approval'
            approval.user = request.user
            approval.save()  # Ensure changes are saved

    # Step 4: Generate a unique invoice number
    timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
    invoice_number = f"PO-{po.id:05d}-{timestamp}"

    # Step 5: Create the invoice
    invoice = Invoice.objects.create(
        purchase_order=po,
        invoice_number=invoice_number,
        vendor=po.vendor,
        date=timezone.now(),
        user=request.user,
        status='Draft',
        method='Purchase Order',
    )

    # Step 6: Add line items to invoice and calculate totals
    sub_total = Decimal('0.00')
    total_sgst = Decimal('0.00')
    total_cgst = Decimal('0.00')
    total_amount = Decimal('0.00')

    for item in po.order_lines.all():
        qty = item.qty
        rate = item.price
        line_total = qty * rate
        sgst = line_total * Decimal('0.09')
        cgst = line_total * Decimal('0.09')
        total_line_amount = line_total + sgst + cgst

        InvoiceLineItem.objects.create(
            invoice=invoice,
            purchase_order_line=item,
            qty=qty,
            amount=rate,
            line_total=line_total,
            sgst_amount=sgst,
            cgst_amount=cgst,
            total_amount=total_line_amount,
            user=request.user,
        )

        sub_total += line_total
        total_sgst += sgst
        total_cgst += cgst
        total_amount += total_line_amount

    # Step 7: Finalize invoice with totals and amount in words
    invoice.sub_total = sub_total
    invoice.total_sgst = total_sgst
    invoice.total_cgst = total_cgst
    invoice.total_amount = total_amount
    invoice.amount_in_words = num2words(total_amount, to='currency', lang='en_IN')
    invoice.save()

    # Step 8: Notify user and redirect
    messages.success(request, "Purchase Order items sent for approval. Invoice generated.")
    return redirect('purchase_order_detail', pk=po.pk)

from django.shortcuts import get_object_or_404, render
from .models import Invoice

def invoice_view(request, pk):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    invoice = get_object_or_404(Invoice, pk=pk)
    return render(request, 'invoice_detail.html', {'invoice': invoice,'bok':bok})


from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from num2words import num2words

from .models import PurchaseOrder, Invoice, Registration


from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from num2words import num2words
from .models import PurchaseOrder, PurchaseOrderApproval, Invoice, Registration

from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from num2words import num2words

from .models import PurchaseOrder, PurchaseOrderApproval, Invoice, Registration


from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from num2words import num2words

def place_purchase_order(request, pk):
    """
    Handles placing of a purchase order, updates approval status, and redirects to invoice.
    """
    po = get_object_or_404(PurchaseOrder, pk=pk)

    if request.method == "POST":
        # Update purchase order status
        po.status = 'placed'
        po.save()

        # Try to fetch related invoice
        invoice = Invoice.objects.filter(purchase_order=po).first()
        if invoice:
            messages.success(request, f"Purchase Order {po.id} has been placed.")
            return redirect('purchase_order_invoice', pk=invoice.pk)

        # If no invoice exists, render invoice preview
        bok_id = request.session.get('logg')
        if not bok_id:
            messages.error(request, "User session not found.")
            return redirect('purchase_order_detail', pk=po.pk)

        bok = get_object_or_404(Registration, id=bok_id)
        order_lines = po.order_lines.all()

        total_incl_gst = Decimal('0.00')
        for line in order_lines:
            rate = Decimal(line.amount or 0)
            qty = Decimal(line.qty or 0)
            gst = Decimal(line.gst or 0)

            gross = rate * qty
            gst_amt = gross * gst / Decimal('100')
            total_incl_gst += gross + gst_amt

        # Convert total amount to words
        amount_in_words = num2words(
            total_incl_gst.quantize(Decimal('0.01')),
            to='currency',
            lang='en_IN'
        ).replace('euro', 'Rupees').replace('cents', 'Only')

       
        return render(request, 'purchase_order_invoice.html', {
            'purchase_order': po,
            'order_lines': order_lines,
            'bok': bok,
            'amount_in_words': amount_in_words,
        })

    return redirect('purchase_order_detail', pk=po.pk)
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404, render
from .models import PurchaseOrder, Invoice, InvoiceLineItem
from num2words import num2words
from datetime import date
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from decimal import Decimal
from datetime import date
from num2words import num2words

from .models import PurchaseOrder, Invoice, InvoiceLineItem


from datetime import date
from django.db.models import Max
from myapp.models import Invoice


def generate_invoice_number():
    """
    Generates a unique invoice number in the format:
    INV-YYYY-0001
    """
    year = date.today().year
    prefix = f"INV-{year}-"

    # Get the last invoice number for the current year
    last_invoice = (
        Invoice.objects
        .filter(invoice_number__startswith=prefix)
        .aggregate(last_no=Max("invoice_number"))
        .get("last_no")
    )

    if last_invoice:
        try:
            last_number = int(last_invoice.split("-")[-1])
        except (ValueError, IndexError):
            last_number = 0
    else:
        last_number = 0

    next_number = last_number + 1

    return f"{prefix}{next_number:04d}"



from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from datetime import date
from num2words import num2words
from .models import PurchaseOrder, Invoice, InvoiceLineItem


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal
from datetime import date
from num2words import num2words
from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval

from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from datetime import date
from num2words import num2words

from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval

from decimal import Decimal, InvalidOperation
from datetime import date
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

# Make sure generate_next_po_number() is imported or defined here

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from datetime import date
from num2words import num2words
from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval, PurchaseOrderLine, Vendor, Item, User  # Make sure to import all your models

# (Assuming these functions exist in your code)
# from .utils import generate_next_po_number 





def generate_next_po_number():
    """Generates the next PO number in the format PO-000001."""
    latest_po = PurchaseOrder.objects.order_by('-id').first()
    if latest_po and latest_po.invoice_number and latest_po.invoice_number.startswith("PO-"):
        try:
            # Extract number part and increment
            last_number = int(latest_po.invoice_number.split('-')[-1])
            return f"PO-{last_number + 1:06d}"
        except ValueError:
            pass
    # Default first PO number in correct format
    return "PO-000001"

from decimal import Decimal, InvalidOperation
from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect

from num2words import num2words

from .models import (
    PurchaseOrder,
    Invoice,
    InvoiceLineItem,
    PurchaseOrderApproval
)





from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.shortcuts import redirect
from django.contrib import messages
from .models import PurchaseOrder, PurchaseOrderLine, Vendor, Item

def create_purchase_order(request):
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('company_list')

    try:
        vendor_id = request.POST.get('vendor_id')
        if not vendor_id or not vendor_id.isdigit():
            messages.error(request, "Invalid Vendor ID.")
            return redirect('company_list')

        vendor = Vendor.objects.get(id=int(vendor_id))
        user = request.user if request.user.is_authenticated else None

        # Auto-generate standardized PO number
        invoice_number = generate_next_po_number()

        # Handle date field (default to today if not provided)
        date_str = request.POST.get('dateField')
        if date_str:
            try:
                order_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
                return redirect('company_list')
        else:
            order_date = datetime.today().date()

        # Create Purchase Order
        po = PurchaseOrder.objects.create(
            vendor=vendor,
            company=request.POST.get('company', '').strip(),
            invoice_number=invoice_number,
            user=user,
            dateField=order_date,  # ✅ Added date field
        )


        # Fetch item line inputs
        item_ids = request.POST.getlist('item_id[]')
        prices = request.POST.getlist('price[]')
        packs = request.POST.getlist('pack[]')
        stocks = request.POST.getlist('stock[]')
        qtys = request.POST.getlist('qty[]')
        hsn_codes = request.POST.getlist('hsn[]')
        gst_vals = request.POST.getlist('gst[]')

        line_count = len(item_ids)
        if not all(len(lst) == line_count for lst in [prices, packs, stocks, qtys, hsn_codes, gst_vals]):
            messages.error(request, "Mismatch in item line data lengths.")
            return redirect('company_list')

        for i in range(line_count):
            try:
                item = Item.objects.get(id=item_ids[i])
                price = Decimal(prices[i])
                qty = Decimal(qtys[i])
                gst = Decimal(gst_vals[i])
                amount = price * qty
                total_amount = amount + (amount * gst / 100)

                PurchaseOrderLine.objects.create(
                    purchase_order=po,
                    item=item,
                    price=price,
                    pack=packs[i],
                    stock=int(float(stocks[i])),
                    qty=qty,
                    gst=gst,
                    hsn=hsn_codes[i],
                    amount=amount,
                    total_amount=total_amount,
                )
            except (Item.DoesNotExist, InvalidOperation, ValueError) as line_error:
                messages.error(request, f"Error processing line {i + 1}: {str(line_error)}")
                return redirect('company_list')
        log_usage(
            request.user,
            'PO_CREATE'
        )
        messages.success(request, f"Purchase Order Created Successfully. PO Number: {invoice_number}")
        return redirect('view_purchase_orders')

    except Exception as e:
        messages.error(request, f"Unexpected error: {str(e)}")
        return redirect('company_list')
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages

def update_purchase_order(request, po_id):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    po = get_object_or_404(PurchaseOrder, id=po_id)

    if request.method == 'POST':
        try:
            # --- Update PO fields ---
            vendor_id = request.POST.get('vendor_id')
            if not vendor_id or not vendor_id.isdigit():
                messages.error(request, "Invalid Vendor ID.")
                return redirect('view_purchase_orders')

            vendor = get_object_or_404(Vendor, id=int(vendor_id))
            po.vendor = vendor
            po.company = request.POST.get('company', '').strip()

            date_str = request.POST.get('dateField')
            if date_str:
                try:
                    po.dateField = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
                    return redirect('view_purchase_orders')
            else:
                po.dateField = datetime.today().date()

            po.user = request.user if request.user.is_authenticated else None
            po.save()

            # --- Handle lines ---
            line_ids = request.POST.getlist('line_id[]')
            item_ids = request.POST.getlist('item_id[]')
            prices = request.POST.getlist('price[]')
            packs = request.POST.getlist('pack[]')
            stocks = request.POST.getlist('stock[]')
            qtys = request.POST.getlist('qty[]')
            hsn_codes = request.POST.getlist('hsn[]')
            gst_vals = request.POST.getlist('gst[]')
            cgst_vals = request.POST.getlist('cgst[]')
            sgst_vals = request.POST.getlist('sgst[]')

            line_count = len(item_ids)
            if not all(len(lst) == line_count for lst in [prices, packs, stocks, qtys, hsn_codes, gst_vals, cgst_vals, sgst_vals]):
                messages.error(request, "Mismatch in line data lengths.")
                return redirect('view_purchase_orders')

            existing_line_ids = set(po.order_lines.values_list('id', flat=True))
            updated_line_ids = set()

            for i in range(line_count):
                try:
                    item = get_object_or_404(Item, id=int(item_ids[i]))
                    price = Decimal(prices[i])
                    qty = Decimal(qtys[i])
                    gst = Decimal(gst_vals[i] or 0)
                    cgst = Decimal(cgst_vals[i] or 0)
                    sgst = Decimal(sgst_vals[i] or 0)
                    amount = price * qty
                    total_amount = amount + (amount * gst / 100)

                    line_id = line_ids[i] if i < len(line_ids) else None

                    if line_id and line_id.isdigit():
                        pol = get_object_or_404(PurchaseOrderLine, id=int(line_id), purchase_order=po)
                        pol.item = item
                        pol.price = price
                        pol.pack = packs[i]
                        pol.stock = int(float(stocks[i]))
                        pol.qty = qty
                        pol.gst = gst
                        pol.cgst = cgst
                        pol.sgst = sgst
                        pol.hsn = hsn_codes[i]
                        pol.amount = amount
                        pol.total_amount = total_amount
                        pol.save()
                        updated_line_ids.add(pol.id)
                    else:
                        pol = PurchaseOrderLine.objects.create(
                            purchase_order=po,
                            item=item,
                            price=price,
                            pack=packs[i],
                            stock=int(float(stocks[i])),
                            qty=qty,
                            gst=gst,
                            cgst=cgst,
                            sgst=sgst,
                            hsn=hsn_codes[i],
                            amount=amount,
                            total_amount=total_amount,
                        )
                        updated_line_ids.add(pol.id)

                except (Item.DoesNotExist, InvalidOperation, ValueError) as e:
                    messages.error(request, f"Error processing line {i+1}: {str(e)}")
                    return redirect('view_purchase_orders')

            # Delete removed lines
            lines_to_delete = existing_line_ids - updated_line_ids
            if lines_to_delete:
                po.order_lines.filter(id__in=lines_to_delete).delete()

            messages.success(request, f"Purchase Order Updated Successfully. PO Number: {po.invoice_number}")
            return redirect('view_purchase_orders')

        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")
            return redirect('view_purchase_orders')

    # --- GET request ---
    vendors = Vendor.objects.all()
    items = Item.objects.all()
    return render(request, 'update_purchase_order.html', {
        'po': po,
        'vendors': vendors,
        'items': items,
        'bok':bok,
    })


from decimal import Decimal, InvalidOperation
from datetime import date
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from num2words import num2words
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from datetime import date
from num2words import num2words
from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval

@login_required
@transaction.atomic
def create_invoice_from_po(request, pk):

    if request.method != "POST":
        messages.error(request, "Invalid request")
        return redirect("purchase_order_detail", pk=pk)

    po = get_object_or_404(PurchaseOrder, pk=pk)

    if po.is_confirmed:
        messages.warning(request, "Invoice already created")
        return redirect("purchase_order_detail", pk=pk)

    invoice = Invoice.objects.create(
        purchase_order=po,
        invoice_number=generate_invoice_number(),  # ✅ UNIQUE
        vendor=po.vendor,
        user=request.user,
        dateField=po.dateField
    )

    sub_total = total_cgst = total_sgst = total_amount = Decimal("0.00")

    for line in po.order_lines.all():

        gross = line.price * line.qty
        gst_amount = gross * line.gst / 100
        cgst = gst_amount / 2
        sgst = gst_amount / 2
        total = gross + gst_amount

        line.cgst = cgst
        line.sgst = sgst
        line.save()

        InvoiceLineItem.objects.create(
            invoice=invoice,
            purchase_order_line=line,
            qty=line.qty,
            rate=line.price,
            line_total=gross,
            cgst_amount=cgst,
            sgst_amount=sgst,
            total_amount=total,
            user=request.user
        )

        PurchaseOrderApproval.objects.update_or_create(
            purchase_order_line=line,
            defaults={
                "approval_status": True,
                "user": request.user
            }
        )

        sub_total += gross
        total_cgst += cgst
        total_sgst += sgst
        total_amount += total

    invoice.sub_total = sub_total
    invoice.total_cgst = total_cgst
    invoice.total_sgst = total_sgst
    invoice.total_amount = total_amount
    invoice.amount_in_words = num2words(total_amount, to="currency", lang="en_IN")
    invoice.save()

    po.is_confirmed = True
    po.status = "placed"
    po.save()

    messages.success(request, "Invoice created successfully")
    return redirect("purchase_order_detail", pk=pk)
from django.db import transaction
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from num2words import num2words
from .models import PurchaseOrder, Invoice, InvoiceLineItem, PurchaseOrderApproval

from decimal import Decimal
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

from decimal import Decimal
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required

@login_required
@transaction.atomic
def undo_invoice_from_po(request, pk):
    """
    Undo the invoice created from a Purchase Order:
    - Revert approvals
    - Delete invoice line items and adjust Item stock_balance
    - Delete invoice itself
    - Reset PO status
    All actions are atomic.
    """
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('purchase_order_detail', pk=pk)

    po = get_object_or_404(PurchaseOrder, pk=pk)

    # Find the invoice created from this PO
    invoice = Invoice.objects.filter(purchase_order=po).first()
    if not invoice:
        messages.warning(request, "No invoice found for this Purchase Order.")
        return redirect('purchase_order_detail', pk=po.pk)

    try:
        # Reverse approvals for all order lines
        PurchaseOrderApproval.objects.filter(
            purchase_order_line__in=po.order_lines.all()
        ).update(approval_status=False)

        # Adjust Item stock_balance before deleting invoice line items
        invoice_line_items = InvoiceLineItem.objects.filter(invoice=invoice)
        for line in invoice_line_items:
            if line.purchase_order_line and line.purchase_order_line.item:
                item = line.purchase_order_line.item
                # Since we're undoing the invoice, we **add back the qty** to stock
                item.stock_balance = (item.stock_balance or Decimal('0.00')) - (line.qty or Decimal('0.00'))
                item.save()

        # Delete all invoice line items
        invoice_line_items.delete()

        # Delete the invoice
        invoice_number = invoice.invoice_number
        invoice.delete()

        # Reset Purchase Order status
        po.is_confirmed = False
        po.status = "pending"  # default status
        po.save()

        messages.success(
            request, f"Invoice {invoice_number} has been successfully undone and deleted."
        )
        return redirect('purchase_order_detail', pk=po.pk)

    except Exception as e:
        # Any error will rollback everything
        messages.error(request, f"An error occurred while undoing the invoice: {str(e)}")
        raise  # triggers rollback because of @transaction.atomic



@login_required
def view_invoices(request):
    invoices = Invoice.objects.select_related('purchase_order', 'vendor', 'user').order_by('-date')
    return render(request, 'invoice_list.html', {'invoices': invoices})




from django.shortcuts import render, get_object_or_404
from .models import Invoice

def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    return render(request, 'display_purchase_orders.html', {'invoice': invoice})




from django.shortcuts import render, get_object_or_404
from .models import PurchaseOrder

from django.shortcuts import render, get_object_or_404
from .models import PurchaseOrder, Invoice, Registration

from django.shortcuts import render, get_object_or_404
from .models import PurchaseOrder, Invoice, Registration

def purchase_order_detail(request, pk):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    po = get_object_or_404(PurchaseOrder, pk=pk)

    # Get the first related invoice (if it exists)
    invoice = Invoice.objects.filter(purchase_order=po).first()

    return render(request, 'purchase_order_detail.html', {
        'purchase_order': po,
        'invoice': invoice,
        'bok': bok
    })





def company_create(request):

    vendors = Vendor.objects.all()
    items = Item.objects.all()

    if request.method == 'POST':

        name = request.POST.get('name')
        item_id = request.POST.get('item')
        vendor_id = request.POST.get('vendor')
        branch_id = request.POST.get('branch')
        item_prices = request.POST.getlist('item_price')
        pack_sizes = request.POST.getlist('pack_size')
        stock_balances = request.POST.getlist('stock_balance')
        quantities = request.POST.getlist('quantity')
        if not all(len(lst) == len(item_prices) for lst in [pack_sizes, stock_balances, quantities]):
            messages.error(request, "Mismatch in input data. Ensure all fields are correctly filled.")
            return redirect('company_create')
        if not name or not item_id or not vendor_id or not branch_id:
            messages.error(request, "All company fields are required. Please fill in all the fields.")
            return redirect('company_create')

        try:

            company = Company(
                name=name,
                item=Item.objects.get(id=item_id),
                vendor=Vendor.objects.get(id=vendor_id),
     
                branch=Branch.objects.get(id=branch_id)  
            )
            company.save()
            for i in range(len(item_prices)):
                try:

                    item_price = Decimal(item_prices[i]) if item_prices[i] else 0
                    stock_balance = int(stock_balances[i]) if stock_balances[i] else 0
                    quantity = int(quantities[i]) if quantities[i] else 0
                    pack_size = pack_sizes[i] if pack_sizes[i] else '0'

                    Company_line.objects.create(
                        company=company,
                        item_price=item_price,
                        pack_size=pack_size,
                        stock_balance=stock_balance,
                        quantity=quantity
                    )

                except ValueError:
                    messages.error(request, f"Invalid data on line {i+1}. Ensure correct values for item_price, stock_balance, and quantity.")
                    continue

            messages.success(request, "Company and lines created successfully.")
            return redirect('company_list')

        except Item.DoesNotExist:
            messages.error(request, "Invalid item selection. Please choose a valid item.")
        except Vendor.DoesNotExist:
            messages.error(request, "Invalid vendor selection. Please choose a valid vendor.")
        except Branch.DoesNotExist:
            messages.error(request, "Invalid branch selection. Please choose a valid branch.")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
        return redirect('company_create') 
    items = Item.objects.all()
    vendors = Vendor.objects.all()
    branches = Branch.objects.all()  
    return render(request, 'company_list.html', {'items': items, 'vendors': vendors, 'branches': branches})


def company_update(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        company.name = request.POST.get('name')
        company.item = Item.objects.get(id=request.POST.get('item'))
        company.vendor = Vendor.objects.get(id=request.POST.get('vendor'))
        company.item_price = request.POST.get('item_price')
        company.pack_size = request.POST.get('pack_size')
        company.stock_balance = request.POST.get('stock_balance')
        company.quantity = request.POST.get('quantity')
        branch_id = request.POST.get('branch') 
        company.branch = Branch.objects.get(id=branch_id) 
        company.save()
        return redirect('company_list')
    items = Item.objects.all()
    vendors = Vendor.objects.all()
    branches = Branch.objects.all() 
    return render(request, 'company_list.html', {'company': company, 'items': items, 'vendors': vendors, 'branches': branches})
# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Company

def company_delete(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        company.delete()
        return redirect('company_list')  # Redirect to company list after deletion
    return render(request, 'company_confirm_delete.html', {'company': company})




from django.contrib.auth.decorators import login_required



@login_required
def vendor_list_admin(request):
    vendors = Vendor.objects.all()
    return render(request, 'vendor_list_admin.html', {'vendors': vendors})


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Vendor
from django.contrib.auth.decorators import login_required




@login_required
def vendor_list(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))

    search = request.GET.get("search", "").strip()
    date_filter = request.GET.get("date_filter", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    per_page = request.GET.get("per_page", "10").strip()
    sort_by = request.GET.get("sort_by", "-id").strip()

    vendors_qs = Vendor.objects.filter(user=request.user)

    if search:
        vendors_qs = vendors_qs.filter(
            Q(name__icontains=search) |
            Q(mobile_number__icontains=search) |
            Q(email__icontains=search) |
            Q(address__icontains=search)
        )

    today = timezone.localdate()

    if date_filter == "daily":
        vendors_qs = vendors_qs.filter(created_at__date=today)

    elif date_filter == "weekly":
        week_start = today - timedelta(days=today.weekday())
        vendors_qs = vendors_qs.filter(
            created_at__date__gte=week_start,
            created_at__date__lte=today
        )

    elif date_filter == "monthly":
        vendors_qs = vendors_qs.filter(
            created_at__year=today.year,
            created_at__month=today.month
        )

    elif date_filter == "yearly":
        vendors_qs = vendors_qs.filter(created_at__year=today.year)

    elif date_filter == "custom":
        if start_date:
            vendors_qs = vendors_qs.filter(created_at__date__gte=start_date)
        if end_date:
            vendors_qs = vendors_qs.filter(created_at__date__lte=end_date)

    if sort_by == "name_asc":
        vendors_qs = vendors_qs.order_by("name")
    elif sort_by == "name_desc":
        vendors_qs = vendors_qs.order_by("-name")
    elif sort_by == "oldest":
        vendors_qs = vendors_qs.order_by("id")
    else:
        vendors_qs = vendors_qs.order_by("-id")

    total_count = vendors_qs.count()

    entries_options = ["5", "10", "25", "50", "100", "all"]

    if per_page not in entries_options:
        per_page = "10"

    if per_page == "all":
        vendors = vendors_qs
        page_obj = None
        showing_count = total_count
    else:
        paginator = Paginator(vendors_qs, int(per_page))
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        vendors = page_obj
        showing_count = page_obj.object_list.count()

    query_params = request.GET.copy()
    query_params.pop("page", None)

    context = {
        "bok": bok,
        "vendors": vendors,
        "page_obj": page_obj,
        "search": search,
        "date_filter": date_filter,
        "start_date": start_date,
        "end_date": end_date,
        "per_page": per_page,
        "sort_by": sort_by,
        "total_count": total_count,
        "showing_count": showing_count,
        "query_params": query_params.urlencode(),
        "entries_options": entries_options,
    }

    return render(request, "vendor_list.html", context)

@login_required
def vendor_import_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            headers = [str(cell.value).strip().lower() for cell in sheet[1]]

            rows_created = 0
            errors = []
            vendors_to_create = []

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = {"user": request.user}

                    for col_idx, cell_value in enumerate(row):
                        header = headers[col_idx]

                        if header in ["sl", "serial"]:
                            continue
                        elif header in ["name", "vendor name"]:
                            data["name"] = str(cell_value or "").strip()
                        elif header in ["mobile", "mobile number"]:
                            data["mobile_number"] = str(cell_value or "").strip()
                        elif header in ["email"]:
                            data["email"] = str(cell_value or "").strip()
                        elif header in ["address"]:
                            data["address"] = str(cell_value or "").strip()
                        else:
                            data[header] = cell_value

                    vendors_to_create.append(Vendor(**data))
                    rows_created += 1
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            Vendor.objects.bulk_create(vendors_to_create)

            if errors:
                return JsonResponse({
                    "status": "partial_success",
                    "message": f"{rows_created} rows imported, but some rows failed.",
                    "errors": errors
                })

            return JsonResponse({
                "status": "success",
                "message": f"{rows_created} vendors imported successfully in Excel order!"
            })

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "No file uploaded"}, status=400)


@login_required
def vendor_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        mobile_number = request.POST.get('mobile_number', '')
        email = request.POST.get('email', '')
        address = request.POST.get('address', '')
        
        Vendor.objects.create(
            name=name,
            mobile_number=mobile_number,
            email=email,
            address=address,
            user=request.user  # assign current user
        )
        return redirect('vendor_list')
    
    return render(request, 'vendor_form.html')


def vendor_update(request, pk):
    bok = Registration.objects.get(id=request.session['logg'])
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile_number = request.POST.get('mobile_number')
        email = request.POST.get('email')
        address = request.POST.get('address')

        if name is not None:
            vendor.name = name
        if mobile_number is not None:
            vendor.mobile_number = mobile_number
        if email is not None:
            vendor.email = email
        if address is not None:
            vendor.address = address

        vendor.save()
        return redirect('vendor_list')
    
    return render(request, 'vendor_update.html', {'vendor': vendor,'bok':bok})


def vendor_delete(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        vendor.delete()
        return redirect('vendor_list')
    return render(request, 'vendor_confirm_delete.html', {'vendor': vendor})

@login_required
@csrf_exempt
def vendor_bulk_delete(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            vendor_ids = data.get('ids', [])
            if vendor_ids:
                Vendor.objects.filter(id__in=vendor_ids, user=request.user).delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)



def vendor_delete_admin(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        vendor.delete()
        return redirect('vendor_list_admin')
    return render(request, 'vendor_confirm_delete.html', {'vendor': vendor})

from django.contrib.auth.decorators import login_required
from .models import Item, Vendor

@login_required
def item_list_admin(request):
    items = Item.objects.all()
    vendors = Vendor.objects.all()
    return render(request, 'item_list_admin.html', {
        'items': items,
        'vendors': vendors
    })


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Item, Vendor


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from .models import Item, Vendor, Registration


@login_required
def item_import_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            headers = [str(cell.value).strip().lower() for cell in sheet[1]]

            rows_created = 0
            errors = []

            # Keep all Item instances in a list for bulk_create to preserve order
            items_to_create = []

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = {"user": request.user}  # attach user for all rows

                    for col_idx, cell_value in enumerate(row):
                        header = headers[col_idx]

                        if header in ["sl", "serial"]:
                            continue
                        elif header in ["item name", "item nam", "name"]:
                            data["name"] = str(cell_value or "").strip()
                        elif header in ["company"]:
                            data["company_name"] = str(cell_value or "").strip()
                        elif header in ["hsn"]:
                            data["hsn"] = str(cell_value or "").strip()
                        elif header in ["vendor"]:
                            vendor_name = str(cell_value or "Unknown").strip()
                            vendor_obj, _ = Vendor.objects.get_or_create(
                                name=vendor_name, user=request.user
                            )
                            data["vendor"] = vendor_obj
                        elif header in ["pack", "pack size"]:
                            data["pack_size"] = str(cell_value or "").strip()
                        elif header in ["gst", "tax"]:
                            try:
                                val = str(cell_value or "0").lower().replace("rs", "").strip()
                                data["gst"] = float(val)
                            except:
                                data["gst"] = 0
                        elif header in ["price", "item price"]:
                            try:
                                val = str(cell_value or "0").lower().replace("rs", "").strip()
                                data["item_price"] = float(val)
                            except:
                                data["item_price"] = 0
                        elif header in ["stock", "stock balance"]:
                            try:
                                data["stock_balance"] = float(cell_value or 0)
                            except:
                                data["stock_balance"] = 0
                        else:
                            data[header] = cell_value

                    # Instead of creating one by one, prepare objects to preserve order
                    items_to_create.append(Item(**data))
                    rows_created += 1

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            # Bulk create items in the exact order of Excel rows
            Item.objects.bulk_create(items_to_create)

            if errors:
                return JsonResponse({
                    "status": "partial_success",
                    "message": f"{rows_created} rows imported, but some rows failed.",
                    "errors": errors
                })

            return JsonResponse({
                "status": "success",
                "message": f"{rows_created} items imported successfully in Excel order!"
            })

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "No file uploaded"}, status=400)


@login_required
def item_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        vendor_id = request.POST.get('vendor')
        item_price = request.POST.get('item_price')
        pack_size = request.POST.get('pack_size')
        stock_balance = request.POST.get('stock_balance')
        gst = request.POST.get('gst')
        hsn = request.POST.get('hsn')
        company_name = request.POST.get('company_name')

        vendor = Vendor.objects.get(id=vendor_id)

        Item.objects.create(
            name=name,
            vendor=vendor,
            item_price=item_price,
            pack_size=pack_size,
            stock_balance=stock_balance,
            gst=gst,
            hsn=hsn,
            company_name=company_name,
            user=request.user  # Set logged-in user
        )

        return redirect('item_list')

    vendors = Vendor.objects.all()
    return render(request, 'item_form.html', {'vendors': vendors})


from django.urls import reverse

def item_update(request, pk):
    bok = Registration.objects.get(id=request.session['logg'])
    item = get_object_or_404(Item, pk=pk)

    if request.method == 'POST':
        item.name = request.POST.get('name')
        item.vendor = Vendor.objects.get(id=request.POST.get('vendor'))
        item.item_price = request.POST.get('item_price')
        item.pack_size = request.POST.get('pack_size')
        item.gst = request.POST.get('gst')
        item.hsn = request.POST.get('hsn')
        item.stock_balance = request.POST.get('stock_balance')
        item.company_name = request.POST.get('company_name')
        item.save()
        return redirect(f"{reverse('item_list')}?highlight={item.id}")

    vendors = Vendor.objects.all()
    return render(request, 'update_item.html', {'item': item, 'vendors': vendors,'bok':bok})



@login_required
def item_list(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))

    search = request.GET.get("search", "").strip()
    date_filter = request.GET.get("date_filter", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    per_page = request.GET.get("per_page", "10").strip()
    sort_by = request.GET.get("sort_by", "-id").strip()

    items_qs = Item.objects.select_related("vendor").filter(
        Q(user=request.user) | Q(user__isnull=True)
    )

    vendors = Vendor.objects.filter(
        Q(user=request.user) | Q(user__isnull=True)
    )

    if search:
        items_qs = items_qs.filter(
            Q(name__icontains=search) |
            Q(company_name__icontains=search) |
            Q(hsn__icontains=search) |
            Q(vendor__name__icontains=search) |
            Q(pack_size__icontains=search) |
            Q(gst__icontains=search) |
            Q(item_price__icontains=search) |
            Q(stock_balance__icontains=search)
        )

    today = timezone.localdate()

    if date_filter == "daily":
        items_qs = items_qs.filter(created_at=today)

    elif date_filter == "weekly":
        week_start = today - timedelta(days=today.weekday())
        items_qs = items_qs.filter(
            created_at__gte=week_start,
            created_at__lte=today
        )

    elif date_filter == "monthly":
        items_qs = items_qs.filter(
            created_at__year=today.year,
            created_at__month=today.month
        )

    elif date_filter == "yearly":
        items_qs = items_qs.filter(created_at__year=today.year)

    elif date_filter == "custom":
        if start_date:
            items_qs = items_qs.filter(created_at__gte=start_date)
        if end_date:
            items_qs = items_qs.filter(created_at__lte=end_date)

    if sort_by == "name_asc":
        items_qs = items_qs.order_by("name")
    elif sort_by == "name_desc":
        items_qs = items_qs.order_by("-name")
    elif sort_by == "oldest":
        items_qs = items_qs.order_by("id")
    else:
        items_qs = items_qs.order_by("-id")

    total_count = items_qs.count()

    entries_options = ["5", "10", "25", "50", "100", "all"]

    if per_page not in entries_options:
        per_page = "10"

    if per_page == "all":
        items = items_qs
        page_obj = None
        showing_count = total_count
    else:
        paginator = Paginator(items_qs, int(per_page))
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        items = page_obj
        showing_count = page_obj.object_list.count()

    query_params = request.GET.copy()
    query_params.pop("page", None)

    context = {
        "bok": bok,
        "items": items,
        "vendors": vendors,
        "page_obj": page_obj,
        "search": search,
        "date_filter": date_filter,
        "start_date": start_date,
        "end_date": end_date,
        "per_page": per_page,
        "sort_by": sort_by,
        "total_count": total_count,
        "showing_count": showing_count,
        "query_params": query_params.urlencode(),
        "entries_options": entries_options,
    }

    return render(request, "item_list.html", context)

import json
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST

import json

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST


import json

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST


@login_required
@require_POST
def item_bulk_delete(request):
    try:
        ids = []

        if request.content_type == "application/json":
            data = json.loads(request.body.decode("utf-8") or "{}")
            ids = data.get("ids", [])
        else:
            ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")

        ids = [int(item_id) for item_id in ids if str(item_id).isdigit()]

        if not ids:
            return JsonResponse({
                "success": False,
                "message": "No valid items selected.",
                "deleted_count": 0
            }, status=400)

        items_qs = Item.objects.filter(
            Q(user=request.user) | Q(user__isnull=True),
            id__in=ids
        )

        deleted_item_count = items_qs.count()

        if deleted_item_count == 0:
            return JsonResponse({
                "success": False,
                "message": "Selected items not found or you do not have permission to delete them.",
                "deleted_count": 0
            }, status=404)

        items_qs.delete()

        return JsonResponse({
            "success": True,
            "message": f"{deleted_item_count} item(s) deleted successfully.",
            "deleted_count": deleted_item_count
        })

    except json.JSONDecodeError:
        return JsonResponse({
            "success": False,
            "message": "Invalid JSON data.",
            "deleted_count": 0
        }, status=400)

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": str(e),
            "deleted_count": 0
        }, status=500)

def item_delete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('item_list')
    return render(request, 'item_confirm_delete.html', {'item': item})

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

@login_required
@csrf_exempt  # Needed if using fetch with JSON
def item_bulk_delete(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_ids = data.get('ids', [])
            if item_ids:
                Item.objects.filter(id__in=item_ids).delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

import pandas as pd
from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from .models import Item, Vendor
from django.shortcuts import render
from django.http import JsonResponse
from .models import Item, Vendor
import openpyxl




def item_delete_admin(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('item_list_admin')
    return render(request, 'item_confirm_delete.html', {'item': item})



import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def download_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = ['Company', 'Item', 'Price', 'Stock', 'Pack', 'Status', 'Method']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    for row_num, item in enumerate(PurchaseOrderLine.objects.all(), 2):
        ws[f"A{row_num}"] = item.company
        ws[f"B{row_num}"] = item.item.name
        ws[f"C{row_num}"] = item.price
        ws[f"D{row_num}"] = item.stock
        ws[f"E{row_num}"] = item.pack
        ws[f"F{row_num}"] = item.status if hasattr(item, 'status') else 'Pending'
        ws[f"G{row_num}"] = item.method if hasattr(item, 'method') else 'N/A'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=PurchaseOrders.xlsx'
    wb.save(response)
    return response
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from django.core.mail import EmailMessage
from decimal import Decimal
from num2words import num2words
from xhtml2pdf import pisa
from io import BytesIO
import os
import urllib.parse

import urllib.parse
import os
from django.http import HttpResponse
from django.core.mail import EmailMessage
from decimal import Decimal
from num2words import num2words
from io import BytesIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.shortcuts import get_object_or_404, render
from django.conf import settings
import os
import urllib.parse
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from num2words import num2words
from io import BytesIO
from xhtml2pdf import pisa

from decimal import Decimal, ROUND_HALF_UP

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, FileResponse, Http404
from django.core.mail import EmailMessage
from django.template.loader import get_template
from io import BytesIO
from django.urls import reverse
import os
import urllib.parse

# For PDF generation with xhtml2pdf
from xhtml2pdf import pisa


# Your Django settings import
from django.conf import settings

from num2words import num2words


from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import os
import urllib.parse
from num2words import num2words
from django.conf import settings

from datetime import datetime, timedelta
from django.shortcuts import render


from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from django.shortcuts import render


@login_required
def view_po_items(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    vendor_name = request.GET.get('vendor', '').strip()
    branch_location = request.GET.get('branch', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    date_range = request.GET.get('date_range', 'daily').strip()

    # Filter items by the logged-in user
    orders = PurchaseOrderLine.objects.select_related('vendor', 'branch') \
        .filter(user=request.user) \
        .order_by('-created_at')

    if vendor_name:
        orders = orders.filter(vendor__name__icontains=vendor_name)

    if branch_location:
        orders = orders.filter(branch__location__icontains=branch_location)

    if status:
        orders = orders.filter(status__iexact=status)  # Only if your model has a 'status' field

    today = datetime.today().date()

    if date_range == 'daily':
        orders = orders.filter(created_at=today)

    elif date_range == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        orders = orders.filter(created_at__gte=week_start)

    elif date_range == 'monthly':
        month_start = today.replace(day=1)
        orders = orders.filter(created_at__gte=month_start)

    elif date_range == 'custom':
        if start_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                orders = orders.filter(created_at__gte=start)
            except ValueError:
                pass

        if end_date:
            try:
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                orders = orders.filter(created_at__lte=end)
            except ValueError:
                pass

    # Optional: restrict dropdowns to user's vendors only
    vendors = Vendor.objects.filter(user=request.user)
    branches = Branch.objects.all()  # Filter if branch is user-specific

    context = {
        'orders': orders,
        'vendors': vendors,
        'branches': branches,
        'today': today,
        'bok':bok,
    }

    return render(request, 'view_po.html', context)

@login_required
def view_po_items_admin(request):
    vendor_name = request.GET.get('vendor', '').strip()
    branch_location = request.GET.get('branch', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    date_range = request.GET.get('date_range', 'daily').strip()

    today = datetime.today().date()

    orders = PurchaseOrderLine.objects.select_related('purchase_order', 'item__vendor') \
        .order_by('-purchase_order__created_at')

    if vendor_name:
        orders = orders.filter(item__vendor__name__icontains=vendor_name)

    # Remove this if 'branch' isn't in PurchaseOrder
    if branch_location:
        orders = orders.filter(purchase_order__branch__location__icontains=branch_location)

    if status:
        orders = orders.filter(purchase_order__status__iexact=status)

    if date_range == 'daily':
        orders = orders.filter(purchase_order__created_at=today)

    elif date_range == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        orders = orders.filter(purchase_order__created_at__gte=week_start)

    elif date_range == 'monthly':
        month_start = today.replace(day=1)
        orders = orders.filter(purchase_order__created_at__gte=month_start)

    elif date_range == 'custom':
        if start_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                orders = orders.filter(purchase_order__created_at__gte=start)
            except ValueError:
                pass

        if end_date:
            try:
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                orders = orders.filter(purchase_order__created_at__lte=end)
            except ValueError:
                pass

    vendors = Vendor.objects.filter(user=request.user)
    branches = Branch.objects.all()

    context = {
        'orders': orders,
        'vendors': vendors,
        'branches': branches,
        'today': today,
    }

    return render(request, 'view_po_admin.html', context)

from datetime import datetime
from django.shortcuts import render


from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from django.shortcuts import render

from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta

from django.shortcuts import render
from .models import Invoice

from django.shortcuts import render
from django.db.models import Q
from .models import Invoice, Vendor, Branch
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.utils.timezone import datetime
from django.shortcuts import render
from .models import Invoice, Vendor, Branch

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.utils.dateparse import parse_date
from datetime import datetime
from .models import Invoice, Vendor, Branch, Registration

from datetime import datetime, timedelta, date

from datetime import datetime, timedelta, date
from django.db.models import DateTimeField
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils.dateparse import parse_date
from datetime import datetime, timedelta, date
from django.db.models import DateTimeField

from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta, datetime
from django.db.models import Q
from django.db.models.fields import DateTimeField
from .models import Invoice, Vendor, Branch, Registration

from datetime import date, datetime, timedelta
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Invoice, Vendor, Branch, Registration

from datetime import date, datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404

from datetime import date, datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from datetime import date, timedelta, datetime

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, render

@login_required
def placed_orders(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))

    vendor_name = request.GET.get('vendor', '').strip()
    status = request.GET.get('status', '').strip()
    date_range = request.GET.get('date_range', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    invoice_filter = request.GET.get('invoice_filter', '').strip()

    per_page = request.GET.get('per_page', '10').strip()

    if not per_page:
        per_page = '10'

    invoices = (
        Invoice.objects
        .select_related('vendor', 'purchase_order')
        .prefetch_related('line_items__purchase_order_line__item')
        .filter(user=request.user)
        .order_by('-created_at')
    )

    if vendor_name:
        invoices = invoices.filter(vendor__name__icontains=vendor_name)

    if status:
        invoices = invoices.filter(status__iexact=status)

    if invoice_filter == "invoice":
        invoices = invoices.exclude(status__iexact="Draft")

    elif invoice_filter == "waiting":
        invoices = invoices.filter(status__iexact="Draft")

    today = date.today()
    field_lookup = "invoice_date"

    if date_range == "daily":
        invoices = invoices.filter(**{field_lookup: today})

    elif date_range == "weekly":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        invoices = invoices.filter(**{f"{field_lookup}__range": [start, end]})

    elif date_range == "monthly":
        start = today.replace(day=1)

        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = start.replace(month=start.month + 1, day=1) - timedelta(days=1)

        invoices = invoices.filter(**{f"{field_lookup}__range": [start, end]})

    elif date_range == "custom":
        if start_date:
            invoices = invoices.filter(
                **{f"{field_lookup}__gte": datetime.strptime(start_date, "%Y-%m-%d").date()}
            )

        if end_date:
            invoices = invoices.filter(
                **{f"{field_lookup}__lte": datetime.strptime(end_date, "%Y-%m-%d").date()}
            )

    invoices = invoices.distinct()

    if per_page == "all":
        paginated_invoices = invoices
    else:
        try:
            per_page_value = int(per_page)
        except ValueError:
            per_page_value = 10

        paginator = Paginator(invoices, per_page_value)
        page_number = request.GET.get("page")
        paginated_invoices = paginator.get_page(page_number)

    context = {
        'invoices': paginated_invoices,
        'vendors': Vendor.objects.all(),
        'branches': Branch.objects.all(),
        'bok': bok,
    }

    return render(request, 'placed_orders.html', context)




from django.contrib.auth.models import User  # import if not already

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.dateparse import parse_date
from datetime import datetime
from .models import Invoice, Vendor, Registration  # Registration used instead of direct User if needed
from django.contrib.auth.models import User

@login_required
def placed_orders_all(request):
    vendor_name = request.GET.get('vendor', '').strip()
    user_id = request.GET.get('user', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()

    invoices = Invoice.objects.select_related('vendor', 'user') \
        .prefetch_related('line_items__purchase_order_line__item') \
        .order_by('-created_at')

    if vendor_name:
        invoices = invoices.filter(vendor__name__icontains=vendor_name)

    if user_id:
        invoices = invoices.filter(user__id=user_id)

    if status:
        invoices = invoices.filter(status__iexact=status)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            invoices = invoices.filter(created_at__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            invoices = invoices.filter(created_at__lte=end)
        except ValueError:
            pass

    invoices = invoices.distinct()

    vendors = Vendor.objects.all()
    users = User.objects.all()

    context = {
        'invoices': invoices,
        'vendors': vendors,
        'users': users,
    }

    return render(request, 'placed_orders_all.html', context)


from django.views.decorators.http import require_POST
from django.shortcuts import redirect
from .models import Invoice
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.db import transaction
from .models import Invoice, PurchaseOrderApproval
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.db import transaction
from .models import Invoice, PurchaseOrderApproval
from django.views.decorators.http import require_POST
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from decimal import Decimal
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect
from .models import Invoice, PurchaseOrderApproval

from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from decimal import Decimal
from .models import Invoice, PurchaseOrderApproval

from decimal import Decimal
from django.shortcuts import redirect, get_object_or_404
from django.db import transaction
from django.contrib import messages
from django.views.decorators.http import require_POST
import logging

logger = logging.getLogger(__name__)

from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.db import transaction
from django.contrib import messages
import logging
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from decimal import Decimal
from .models import Invoice, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval
from django.views.decorators.http import require_POST

@require_POST
def invoice_delete_admin(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    with transaction.atomic():
        # Get all related line items before deleting
        line_items = list(invoice.line_items.all())

        for line_item in line_items:
            purchase_order_item = line_item.purchase_order_item
            item = getattr(purchase_order_item, 'item', None)

            # ✅ Restore stock balance
            if item and item.stock_balance is not None:
                current_stock = Decimal(item.stock_balance or 0)
                qty_to_restore = Decimal(line_item.qty or 0)
                item.stock_balance = current_stock - int(qty_to_restore)
                item.save()

            # ✅ Reset approval status
            try:
                approval = PurchaseOrderApproval.objects.get(purchase_order_item=purchase_order_item)
                approval.approval_status = 'Place Order'
                approval.save()
            except PurchaseOrderApproval.DoesNotExist:
                pass

            # ✅ Delete the line item
            line_item.delete()

        # ✅ Delete the invoice itself
        invoice.delete()

    return redirect('placed_orders')

from django.views.decorators.http import require_POST
from django.shortcuts import redirect
from django.db import transaction
from decimal import Decimal


from decimal import Decimal
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from .models import Invoice, PurchaseOrderApproval

@require_POST
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    with transaction.atomic():
        line_items = list(invoice.line_items.all())

        for line_item in line_items:
            purchase_order_line = line_item.purchase_order_line
            item = getattr(purchase_order_line, 'item', None)

            # Restore stock balance
            if item and item.stock_balance is not None:
                current_stock = Decimal(item.stock_balance or 0)
                qty_to_restore = Decimal(line_item.qty or 0)
                item.stock_balance = current_stock - qty_to_restore
                item.save()

            # Reset approval status
            PurchaseOrderApproval.objects.filter(
                purchase_order_line=purchase_order_line
            ).update(approval_status=False)

            # Delete line item
            line_item.delete()

        # Finally delete invoice
        invoice.delete()

    return redirect('placed_orders')

@require_POST
def invoice_bulk_delete(request):
    ids = request.POST.get('selected_ids', '')

    if ids:
        id_list = [int(pk) for pk in ids.split(',') if pk.isdigit()]
        invoices = Invoice.objects.filter(pk__in=id_list)

        with transaction.atomic():
            for invoice in invoices:
                line_items = list(invoice.line_items.all())

                for line_item in line_items:
                    purchase_order_line = line_item.purchase_order_line
                    item = getattr(purchase_order_line, 'item', None)

                    # Restore stock balance
                    if item and item.stock_balance is not None:
                        current_stock = Decimal(item.stock_balance or 0)
                        qty_to_restore = Decimal(line_item.qty or 0)
                        item.stock_balance = current_stock - qty_to_restore
                        item.save()

                    # Reset approval status to False
                    PurchaseOrderApproval.objects.filter(
                        purchase_order_line=purchase_order_line
                    ).update(approval_status=False)

                    line_item.delete()

                invoice.delete()

    return redirect('placed_orders')



from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from .models import Invoice, InvoiceLineItem, PurchaseOrderApproval

from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.db import transaction
from .models import Invoice, PurchaseOrderApproval


from django.shortcuts import get_object_or_404, redirect
from .models import Invoice


# views.py
from django.shortcuts import get_object_or_404, render
from .models import Invoice




    
from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.urls import reverse
import os
import urllib.parse
from num2words import num2words


from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import urllib.parse
import os
from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import os
import urllib.parse
from num2words import num2words
from django.conf import settings

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import os
import urllib.parse
from num2words import num2words
from django.conf import settings

from django.shortcuts import get_object_or_404, redirect
from .models import Invoice, PurchaseOrderApproval

from django.shortcuts import get_object_or_404, redirect
from .models import Invoice, PurchaseOrderApproval
from django.db import transaction



from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import urllib.parse
from num2words import num2words

from django.conf import settings
import os
from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.urls import reverse
import urllib.parse
import os
from num2words import num2words

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.conf import settings
from django.core.mail import EmailMessage
from django.urls import reverse
import os
import urllib.parse
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words



from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
from django.conf import settings
import os
import urllib.parse
from num2words import num2words

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.urls import reverse
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
import urllib



from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import urllib.parse
from num2words import num2words

from django.conf import settings
from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
from django.conf import settings
from num2words import num2words
import urllib.parse

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.conf import settings
from num2words import num2words
from .models import Invoice, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval, Registration
from django.shortcuts import render, get_object_or_404, redirect
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
from .models import Invoice, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval

from django.shortcuts import get_object_or_404, render, redirect
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
from .models import Invoice, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval
from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from num2words import num2words
from .models import Registration, Invoice, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from num2words import num2words
from decimal import Decimal
from datetime import datetime, date
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from num2words import num2words
from .models import Registration, Invoice, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval

from decimal import Decimal
from datetime import datetime
from num2words import num2words

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Invoice, PurchaseOrder, InvoiceLineItem, PurchaseOrderLine, PurchaseOrderApproval, Vendor, Registration
from django.urls import reverse
from django.shortcuts import redirect
from django.core.paginator import Paginator
from datetime import datetime
from decimal import Decimal

@login_required
def invoice_edit(request, pk):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    invoice = get_object_or_404(Invoice, pk=pk)
    vendor = invoice.vendor

    if request.method == 'POST':
        invoice_number = request.POST.get('invoice_number')
        status = request.POST.get('status') or invoice.status
        method = request.POST.get('method') or invoice.method
        invoice_date_str = request.POST.get('invoice_date')

        po_dateField_str = request.POST.get('po_dateField')
        po_created_at_str = request.POST.get('po_created_at')

        if not invoice_number or not invoice_date_str:
            return render(request, 'error.html', {'message': 'Invoice number and date are required.'})

        if Invoice.objects.exclude(pk=invoice.pk).filter(invoice_number=invoice_number).exists():
            return render(request, 'error.html', {'message': 'Invoice number must be unique.'})

        invoice.invoice_number = invoice_number
        invoice.status = status
        invoice.method = method

        try:
            invoice.invoice_date = datetime.strptime(invoice_date_str, "%Y-%m-%d").date()
        except ValueError:
            return render(request, 'error.html', {'message': 'Invalid invoice date format.'})

        invoice.save()

        if invoice.purchase_order:
            po = invoice.purchase_order
            po.invoice_number = invoice_number

            if po_dateField_str:
                try:
                    po.dateField = datetime.strptime(po_dateField_str, "%Y-%m-%d").date()
                except ValueError:
                    pass

            if po_created_at_str:
                try:
                    po.created_at = datetime.strptime(po_created_at_str, "%Y-%m-%d").date()
                except ValueError:
                    pass

            po.save()

        InvoiceLineItem.objects.filter(invoice=invoice).delete()

        sub_total = Decimal('0.00')
        total_sgst = Decimal('0.00')
        total_cgst = Decimal('0.00')

        line_item_ids = request.POST.getlist('line_item_id')

        for line_item_id in line_item_ids:
            try:
                po_line = PurchaseOrderLine.objects.get(
                    id=line_item_id,
                    purchase_order__vendor=vendor
                )
            except PurchaseOrderLine.DoesNotExist:
                continue

            qty = Decimal(request.POST.get(f'qty_{line_item_id}', '0') or '0')
            rate = Decimal(request.POST.get(f'rate_{line_item_id}', '0') or '0')
            gst_rate = Decimal(request.POST.get(f'gst_{line_item_id}', po_line.gst or '0'))

            expiry_date_str = request.POST.get(f'expiry_date_{line_item_id}')
            expiry_date = None

            if expiry_date_str:
                try:
                    expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
                except ValueError:
                    expiry_date = None

            hsn_code = request.POST.get(f'hsn_{line_item_id}', po_line.hsn or '')

            line_total = (rate * qty).quantize(Decimal('0.01'))
            sgst = (line_total * gst_rate / 2 / 100).quantize(Decimal('0.01'))
            cgst = sgst
            total_line_amount = (line_total + sgst + cgst).quantize(Decimal('0.01'))

            InvoiceLineItem.objects.create(
                invoice=invoice,
                purchase_order_line=po_line,
                qty=qty,
                rate=rate,
                line_total=line_total,
                sgst_amount=sgst,
                cgst_amount=cgst,
                expiry_date=expiry_date,
                total_amount=total_line_amount,
                user=request.user
            )

            po_line.qty = qty
            po_line.price = rate
            po_line.gst = gst_rate
            po_line.sgst = sgst
            po_line.cgst = cgst
            po_line.expiry_date = expiry_date
            po_line.hsn = hsn_code
            po_line.total_amount = total_line_amount
            po_line.invoice_number = invoice.invoice_number
            po_line.status = status
            po_line.save()

            if po_line.item and po_line.item.stock_balance is not None:
                po_line.item.stock_balance += int(qty)
                po_line.item.save()

            approval, created = PurchaseOrderApproval.objects.get_or_create(
                purchase_order_line=po_line,
                defaults={'approval_status': False, 'user': request.user}
            )

            if not created:
                approval.approval_status = False
                approval.user = request.user
                approval.save()

            sub_total += line_total
            total_sgst += sgst
            total_cgst += cgst

        grand_total = (sub_total + total_sgst + total_cgst).quantize(Decimal('0.01'))

        invoice.sub_total = sub_total
        invoice.total_sgst = total_sgst
        invoice.total_cgst = total_cgst
        invoice.total_amount = grand_total
        invoice.amount_in_words = f"{num2words(grand_total, lang='en_IN').title()} Only"
        invoice.save()

        per_page = 10

        invoice_ids = list(
            Invoice.objects
            .filter(user=request.user)
            .order_by('-created_at')
            .values_list('id', flat=True)
        )

        try:
            index_position = invoice_ids.index(invoice.id)
        except ValueError:
            index_position = 0

        page_number = (index_position // per_page) + 1

        return redirect(
            f"{reverse('placed_orders')}?page={page_number}&per_page={per_page}&highlight={invoice.id}"
        )

    context = {
        'invoice': invoice,
        'vendor': vendor,
        'bok': bok,
        'line_items': invoice.line_items.all(),
    }

    return render(request, 'invoice_edit.html', context)


def preview_invoice(request, pk):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    invoice = get_object_or_404(Invoice, pk=pk)
    vendor = invoice.vendor
    return render(request, 'invoice_preview.html', {
        'invoice': invoice,
        'vendor': vendor,
        'bok':bok
    })
def preview_invoices(request, pk):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    invoice = get_object_or_404(Invoice, pk=pk)
    vendor = invoice.vendor
    return render(request, 'invoice_previews.html', {
        'invoice': invoice,
        'vendor': vendor,
        'bok':bok
    })
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.conf import settings


from num2words import num2words
import os






from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import os
import urllib.parse
from num2words import num2words
from django.conf import settings

import os
import urllib.parse
from decimal import Decimal, ROUND_HALF_UP
from django.conf import settings
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from num2words import num2words

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
from num2words import num2words
import os
import urllib.parse
from django.conf import settings

def updated_generate_invoice_pdf(context):
    template_path = 'view_invoice.html'
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return result.getvalue()
    return None

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, FileResponse, Http404
from django.core.mail import EmailMessage
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.template.loader import get_template
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
from datetime import date
import urllib.parse
import os
from io import BytesIO
from xhtml2pdf import pisa

def generate_invoice_pdf(context):
    template_path = 'invoice_template.html'
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    return result.getvalue() if not pdf.err else None

@login_required

def generate_invoice(request):
    if request.method != 'POST':
        return render(request, 'invoice_template.html', {'message': 'Invalid request method'})

    try:
        employee = Registration.objects.get(user=request.user, User_role='employee')
    except Registration.DoesNotExist:
        return render(request, 'error.html', {'message': 'You are not registered as an employee.'})

    try:
        vendor_id = int(request.POST.get('vendor_id'))
        status = request.POST.get('status')
        method = request.POST.get('method')
        action = request.POST.get('action')
        date_range = request.POST.get('date_range')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
    except (ValueError, TypeError):
        return render(request, 'error.html', {'message': 'Invalid input data'})

    vendor = get_object_or_404(Vendor, pk=vendor_id)
    order_items = PurchaseOrderLine.objects.filter(vendor=vendor, user=request.user)

    today = datetime.today().date()
    if date_range == 'daily':
        order_items = order_items.filter(created_at=today)
    elif date_range == 'weekly':
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        order_items = order_items.filter(created_at__range=(start_week, end_week))
    elif date_range == 'custom':
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            order_items = order_items.filter(created_at__range=(start, end))
        except ValueError:
            return render(request, 'error.html', {'message': 'Invalid custom date format'})

    if not order_items.exists():
        return render(request, 'error.html', {'message': 'No order items found for this vendor and filters.'})

    invoice_number = f"PO-{vendor.id}-{order_items.first().id}"

    invoice, created = Invoice.objects.get_or_create(
        invoice_number=invoice_number,
        defaults={
            'vendor': vendor,
            'user': request.user,
            'sub_total': Decimal('0.00'),
            'total_sgst': Decimal('0.00'),
            'total_cgst': Decimal('0.00'),
            'total_amount': Decimal('0.00'),
            'amount_in_words': '',
            'status': status,
            'method': method
        }
    )

    if not created and invoice.user != request.user:
        invoice.user = request.user
        invoice.save(update_fields=['user'])

    sub_total = total_sgst = total_cgst = total_incl_gst_subtotal = Decimal('0.00')
    summary_list = []

    for item in order_items:
        rate = Decimal(item.amount or 0)
        qty = Decimal(item.qty or 0)
        gst_percent = Decimal(item.gst or 0)

        gross = (rate * qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        sgst = (gross * (gst_percent / 2) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        cgst = (gross * (gst_percent / 2) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_incl_gst = (gross + sgst + cgst).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        item.gross = gross
        item.total = total_incl_gst
        item.invoice_number = invoice_number
        item.status = status
        item.method = method
        item.save(update_fields=['invoice_number'])

        sub_total += gross
        total_sgst += sgst
        total_cgst += cgst
        total_incl_gst_subtotal += total_incl_gst

        if not InvoiceLineItem.objects.filter(invoice=invoice, purchase_order_item=item).exists():
            InvoiceLineItem.objects.create(
                invoice=invoice,
                purchase_order_item=item,
                qty=qty,
                amount=rate,
                line_total=gross,
                sgst_amount=sgst,
                cgst_amount=cgst,
                total_amount=total_incl_gst,
                user=request.user
            )

        if action in ['download', 'email', 'whatsapp']:
            approval, created = PurchaseOrderApproval.objects.get_or_create(
                purchase_order_item=item,
                defaults={'approval_status': 'Waiting for Approval'}
            )
            if not created:
                approval.approval_status = 'Waiting for Approval'
                approval.save(update_fields=['approval_status'])

        summary_list.append({
            'item_name': item.item.name,
            'invoice_number': item.invoice_number,
            'company': item.company,
            'price': item.price,
            'pack': item.pack,
            'stock': item.stock,
            'qty': item.qty,
            'gst': item.gst,
            'hsn': item.hsn,
            'cgst': cgst,
            'sgst': sgst,
            'amount': item.amount,
            'total_amount': total_incl_gst,
            'created_at': item.created_at,
        })

    grand_total = (sub_total + total_sgst + total_cgst).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    amount_words = f"{num2words(grand_total, lang='en_IN').title()} Only"

    invoice.sub_total = sub_total
    invoice.total_sgst = total_sgst
    invoice.total_cgst = total_cgst
    invoice.total_amount = grand_total
    invoice.amount_in_words = amount_words
    invoice.status = status
    invoice.method = method
    invoice.save()

    summary = {
        'sub_total': sub_total,
        'sgst': total_sgst,
        'cgst': total_cgst,
        'total_incl_gst_subtotal': total_incl_gst_subtotal,
        'total_amount': grand_total,
        'amount_in_words': amount_words,
    }

    context = {
        'vendor': vendor,
        'order_items': order_items,
        'line_items_summary': summary_list,
        'summary': summary,
        'bok': employee,
        'status': status,
        'method': method,
        'invoice_number': invoice_number,
        'invoice': invoice,
        'qr_code_url': '',
        'date_range': date_range,
        'start_date': start_date,
        'end_date': end_date,
    }

    if action == 'filter':
        return render(request, 'invoice_template.html', context)

    pdf = generate_invoice_pdf(context)
    if not pdf:
        return render(request, 'error.html', {'message': 'Failed to generate PDF.'})

    if action == 'download':
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{invoice_number}.pdf"'
        return response

    elif action == 'email':
        if not vendor.email:
            return render(request, 'error.html', {'message': 'Vendor email not found.'})
        email = EmailMessage(
            subject=f"Purchase Order {invoice_number}",
            body="Please find attached your Purchase Order.",
            from_email=settings.EMAIL_HOST_USER,
            to=[vendor.email]
        )
        email.attach(f'{invoice_number}.pdf', pdf, 'application/pdf')
        try:
            email.send()
            return render(request, 'success.html', {'message': 'Invoice emailed successfully.'})
        except Exception as e:
            return render(request, 'error.html', {'message': f'Failed to send email: {e}'})

    elif action == 'whatsapp':
        if not vendor.mobile_number:
            return render(request, 'error.html', {'message': 'Vendor mobile number not found.'})
        file_url = request.build_absolute_uri(reverse('download_pdf', args=[invoice_number]))
        message = f"Hello {vendor.name}, Purchase order from Microlab is ready. Download it here: {file_url}"
        encoded_message = urllib.parse.quote(message)
        whatsapp_url = f"https://wa.me/{vendor.mobile_number}?text={encoded_message}"
        return redirect(whatsapp_url)

    return render(request, 'invoice_template.html', context)


@login_required
def download_pdf(request, invoice_number):
    pdf_path = os.path.join(settings.MEDIA_ROOT, f'{invoice_number}.pdf')
    if os.path.exists(pdf_path):
        return FileResponse(open(pdf_path, 'rb'), content_type='application/pdf')
    raise Http404("PDF not found.")



from django.shortcuts import get_object_or_404

from django.shortcuts import render, get_object_or_404
from decimal import Decimal
from django.shortcuts import render, get_object_or_404

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import os
import urllib.parse
from num2words import num2words
from .models import Invoice, PurchaseOrderApproval, Registration


def generate_invoice_pdf_updated(context):
    template_path = 'view_invoice.html'
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return result.getvalue()
    return None

from django.views.decorators.csrf import csrf_exempt

from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.core.mail import EmailMessage
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
import os
from django.conf import settings
import urllib

from .models import Invoice, PurchaseOrderApproval, Registration

from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.urls import reverse
from .models import Invoice, PurchaseOrderApproval, Registration

import os
import urllib.parse
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.urls import reverse
import urllib
import os

@csrf_exempt
def view_invoice(request, invoice_number):
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    vendor = invoice.vendor
    bok = Registration.objects.filter(User_role='employee').first()
    order_items = invoice.line_items.all()

    sub_total = Decimal('0.00')
    total_sgst = Decimal('0.00')
    total_cgst = Decimal('0.00')
    total_incl_gst_subtotal = Decimal('0.00')
    order_items_data = []

    for item in order_items:
        rate = Decimal(item.amount or 0)
        qty = Decimal(item.qty or 0)
        gst_percent = Decimal(item.purchase_order_item.item.gst or 0)

        gross = (rate * qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        sgst = (gross * (gst_percent / 2) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        cgst = (gross * (gst_percent / 2) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_incl_gst = (gross + sgst + cgst).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        sub_total += gross
        total_sgst += sgst
        total_cgst += cgst
        total_incl_gst_subtotal += total_incl_gst

        order_items_data.append({
            'item': item,
            'rate': rate,
            'qty': qty,
            'gst_percent': gst_percent,
            'gross': gross,
            'total_incl_gst': total_incl_gst,
            'expiry_date': item.expiry_date
        })

    grand_total = (sub_total + total_sgst + total_cgst).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    amount_words = f"{num2words(grand_total, lang='en_IN').title()} Only"

    # Update invoice fields
    invoice.sub_total = sub_total
    invoice.total_sgst = total_sgst
    invoice.total_cgst = total_cgst
    invoice.total_amount = grand_total
    invoice.amount_in_words = amount_words
    invoice.save()

    summary = {
        'sub_total': sub_total,
        'sgst': total_sgst,
        'cgst': total_cgst,
        'total_incl_gst_subtotal': total_incl_gst_subtotal,
        'total_amount': grand_total,
        'amount_in_words': amount_words,
    }

    context = {
        'vendor': vendor,
        'invoice': invoice,
        'summary': summary,
        'bok': bok,
        'order_items_data': order_items_data,
        'qr_code_url': '',
    }

    pdf = generate_invoice_pdf_updated(context)
    if pdf is None:
        return render(request, 'error.html', {'message': 'Failed to generate PDF.'})

    # Send PDF based on action
    action = request.POST.get('action')
    pdf_path = os.path.join(settings.MEDIA_ROOT, f'{invoice_number}.pdf')
    with open(pdf_path, 'wb') as f:
        f.write(pdf)

    if action == 'download':
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{invoice_number}.pdf"'
        return response
    elif action == 'email':
        email = EmailMessage(
            subject=f"Invoice {invoice_number}",
            body="Please find attached your invoice.",
            from_email=settings.EMAIL_HOST_USER,
            to=[vendor.email]
        )
        email.attach(f'{invoice_number}.pdf', pdf, 'application/pdf')
        try:
            email.send()
        except Exception as e:
            return render(request, 'error.html', {'message': f'Failed to send email: {e}'})
        return render(request, 'success.html', {'message': 'Invoice emailed successfully.'})
    elif action == 'whatsapp':
        file_url = request.build_absolute_uri(reverse('download_pdf', args=[invoice_number]))
        msg = f"Hello {vendor.name}, your invoice is ready. Download: {file_url}"
        return redirect(f"https://wa.me/{vendor.mobile_number}?text={urllib.parse.quote(msg)}")

    return render(request, 'view_invoice.html', context)



@csrf_exempt
def view_invoice_admin(request, invoice_number):
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    vendor = invoice.vendor
    bok = Registration.objects.filter(User_role='admin').first()
    order_items = invoice.line_items.all()

    sub_total = Decimal('0.00')
    total_sgst = Decimal('0.00')
    total_cgst = Decimal('0.00')
    total_incl_gst_subtotal = Decimal('0.00')
    order_items_data = []

    for item in order_items:
        rate = Decimal(item.amount or 0)
        qty = Decimal(item.qty or 0)
        gst_percent = Decimal(item.purchase_order_item.item.gst or 0)

        gross = (rate * qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        sgst = (gross * (gst_percent / 2) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        cgst = (gross * (gst_percent / 2) / 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_incl_gst = (gross + sgst + cgst).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        sub_total += gross
        total_sgst += sgst
        total_cgst += cgst
        total_incl_gst_subtotal += total_incl_gst

        order_items_data.append({
            'item': item,
            'rate': rate,
            'qty': qty,
            'gst_percent': gst_percent,
            'gross': gross,
            'total_incl_gst': total_incl_gst,
            'expiry_date': item.expiry_date
        })

    grand_total = (sub_total + total_sgst + total_cgst).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    amount_words = f"{num2words(grand_total, lang='en_IN').title()} Only"

    # Update invoice fields
    invoice.sub_total = sub_total
    invoice.total_sgst = total_sgst
    invoice.total_cgst = total_cgst
    invoice.total_amount = grand_total
    invoice.amount_in_words = amount_words
    invoice.save()

    summary = {
        'sub_total': sub_total,
        'sgst': total_sgst,
        'cgst': total_cgst,
        'total_incl_gst_subtotal': total_incl_gst_subtotal,
        'total_amount': grand_total,
        'amount_in_words': amount_words,
    }

    context = {
        'vendor': vendor,
        'invoice': invoice,
        'summary': summary,
        'bok': bok,
        'order_items_data': order_items_data,
        'qr_code_url': '',
    }

    pdf = generate_invoice_pdf_updated(context)
    if pdf is None:
        return render(request, 'error.html', {'message': 'Failed to generate PDF.'})

    # Send PDF based on action
    action = request.POST.get('action')
    pdf_path = os.path.join(settings.MEDIA_ROOT, f'{invoice_number}.pdf')
    with open(pdf_path, 'wb') as f:
        f.write(pdf)

    if action == 'download':
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{invoice_number}.pdf"'
        return response
    elif action == 'email':
        email = EmailMessage(
            subject=f"Invoice {invoice_number}",
            body="Please find attached your invoice.",
            from_email=settings.EMAIL_HOST_USER,
            to=[vendor.email]
        )
        email.attach(f'{invoice_number}.pdf', pdf, 'application/pdf')
        try:
            email.send()
        except Exception as e:
            return render(request, 'error.html', {'message': f'Failed to send email: {e}'})
        return render(request, 'success.html', {'message': 'Invoice emailed successfully.'})
    elif action == 'whatsapp':
        file_url = request.build_absolute_uri(reverse('download_pdf', args=[invoice_number]))
        msg = f"Hello {vendor.name}, your invoice is ready. Download: {file_url}"
        return redirect(f"https://wa.me/{vendor.mobile_number}?text={urllib.parse.quote(msg)}")

    return render(request, 'view_invoice_admin.html', context)



from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.urls import reverse
from django.core.mail import EmailMessage
from django.conf import settings
from num2words import num2words
import urllib.parse

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from decimal import Decimal, ROUND_HALF_UP
from num2words import num2words
import urllib.parse



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.urls import reverse
from django.conf import settings
from decimal import Decimal, ROUND_HALF_UP
from datetime import date
from num2words import num2words
import urllib.parse
import os

@csrf_exempt
def view_purchase_order(request):
    bok = Registration.objects.filter(User_role='admin').first()
    purchase_orders = PurchaseOrderLine.objects.all()

    vendor_name = request.GET.get('vendor')
    date_range = request.GET.get('date_range')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter by vendor name
    if vendor_name:
        purchase_orders = purchase_orders.filter(vendor__name=vendor_name)

    # Filter by date range
    today = datetime.today().date()
    if date_range == 'daily':
        purchase_orders = purchase_orders.filter(created_at=today)
    elif date_range == 'weekly':
        start_of_week = today - timedelta(days=today.weekday())
        purchase_orders = purchase_orders.filter(created_at__gte=start_of_week)
    elif date_range == 'monthly':
        purchase_orders = purchase_orders.filter(created_at__year=today.year, created_at__month=today.month)
    elif date_range == 'custom':
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                purchase_orders = purchase_orders.filter(created_at__range=(start, end))
            except ValueError:
                pass

    vendors = Vendor.objects.all()
    return render(request, 'filtered_purchase_orders.html', {
        'orders': purchase_orders,
        'vendors': vendors,
        'bok':bok,
    })




def download_pdf(request, invoice_number):
    file_path = os.path.join(settings.MEDIA_ROOT, f'{invoice_number}.pdf')
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=f'{invoice_number}.pdf')
    else:
        raise Http404("Not found")


from django.shortcuts import render
from .models import Invoice
from django.db.models import Q

from datetime import date, timedelta
from django.db.models import Q
from django.shortcuts import render
from .models import Invoice, Vendor

def invoice_list(request):
    # Get filter values
    vendor = request.GET.get('vendor', '').strip()
    branch = request.GET.get('branch', '').strip()
    status = request.GET.get('status', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    search = request.GET.get('search', '').strip()
    date_range = request.GET.get('date_range', '').strip()

    # Handle predefined date ranges
    today = date.today()
    if date_range == 'daily':
        start_date = end_date = today
    elif date_range == 'weekly':
        start_date = today - timedelta(days=today.weekday())  # Monday
        end_date = today
    elif date_range == 'monthly':
        start_date = today.replace(day=1)
        end_date = today

    # Start with all invoices
    invoices = Invoice.objects.all()

    # Apply filters
    if vendor:
        invoices = invoices.filter(vendor__name__icontains=vendor)
    if branch:
        invoices = invoices.filter(method__icontains=branch)  # Assuming 'method' means branch
    if status:
        invoices = invoices.filter(status__iexact=status)
    if start_date:
        invoices = invoices.filter(created_at__date__gte=start_date)
    if end_date:
        invoices = invoices.filter(created_at__date__lte=end_date)
    if search:
        invoices = invoices.filter(
            Q(vendor__name__icontains=search) |
            Q(invoice_number__icontains=search)
        )

    # Context for rendering
    context = {
        'invoices': invoices,
        'vendors': Vendor.objects.all(),
        'branches': Invoice.objects.values_list('method', flat=True).distinct(),
        'request': request  # to access GET parameters in the template
    }
    return render(request, 'placed_orders.html', context)


@login_required
def consumption_list(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))
    companies = Company.objects.filter(user=request.user)
    items = Item.objects.filter(user=request.user)
    vendors = Vendor.objects.filter(user=request.user)
    branches = Branch.objects.all()
    
    return render(request, 'consumption_list.html', {
        'companies': companies,
        'items': items,
        'vendors': vendors,
        'branches': branches,
        'bok':bok,
    })


from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from decimal import Decimal
import logging

import os
from django.conf import settings

from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.conf import settings
from django.http import JsonResponse
import os
import logging



logger = logging.getLogger(__name__)

from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.conf import settings
import os
import logging
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.conf import settings
import os


from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.conf import settings
import os


from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.conf import settings
import os


from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from decimal import Decimal
import os
from django.conf import settings

def delete_consumption_item(request, pk):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

    try:
        # Fetch the consumption item by primary key
        consumption = get_object_or_404(ConsumptionItem, pk=pk)

        # Restore the stock balance for the item
        item = consumption.item
        restored_qty = consumption.qty or Decimal('0.00')
        item.stock_balance = (item.stock_balance or Decimal('0.00')) + restored_qty
        item.save(update_fields=['stock_balance'])

        # Remove associated PDF file if it exists
        pdf_filename = f'consumption_{consumption.id}.pdf'
        pdf_path = os.path.join(settings.MEDIA_ROOT, 'consumption_pdfs', pdf_filename)
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

        # Delete the consumption item
        consumption.delete()

        return redirect('view_consumed')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Error: {str(e)}'}, status=500)
import json
import logging
import traceback
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
logger = logging.getLogger(__name__)
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
import json, logging, traceback
logger = logging.getLogger(__name__)

from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.db import transaction
from decimal import Decimal, InvalidOperation
import json
import logging


logger = logging.getLogger(__name__)

@csrf_exempt
@login_required
def add_consumption(request):
    if request.method != 'POST':
        return render(request, 'consumption_list.html', {
            'status': 'error',
            'message': 'Invalid request method.'
        })

    try:
        data = json.loads(request.body)

        # Required fields
        required_fields = [
            'item_id', 'branch_id', 'company', 'price', 'pack',
            'hsn', 'gst', 'stock', 'qty', 'vendor_id'
        ]
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'status': 'error',
                    'message': f'Missing or empty field: {field}'
                }, status=400)

        # Fetch related objects
        item = get_object_or_404(Item, id=data['item_id'])
        branch = get_object_or_404(Branch, id=data['branch_id'])
        vendor = get_object_or_404(Vendor, id=data['vendor_id'])

        # Convert and validate numeric fields
        try:
            price = Decimal(str(data['price']))
            gst = Decimal(str(data['gst']))
            stock = Decimal(str(data['stock']))
            qty = Decimal(str(data['qty']))
        except (InvalidOperation, TypeError, ValueError):
            return JsonResponse({
                'status': 'error',
                'message': 'Price, GST, stock, and qty must be valid numbers.'
            }, status=400)

        # Business logic
        if item.stock_balance is None:
            current_stock = Decimal('0.00')
        else:
            current_stock = Decimal(item.stock_balance)

        if current_stock < qty:
            return JsonResponse({
                'status': 'error',
                'message': f'Not enough stock available. Current: {current_stock}, Required: {qty}'
            }, status=400)

        # Calculation
        amount = (price * qty).quantize(Decimal('0.01'))
        total_tax = (amount * gst / Decimal('100')).quantize(Decimal('0.01'))
        cgst = sgst = (total_tax / 2).quantize(Decimal('0.01'))
        total_amount = (amount + cgst + sgst).quantize(Decimal('0.01'))

        # Save to DB
        with transaction.atomic():
            item.stock_balance = (current_stock - qty).quantize(Decimal('0.01'))
            item.save()

            consumption = ConsumptionItem.objects.create(
                item=item,
                branch=branch,
                vendor=vendor,
                company=data['company'],
                price=price,
                pack=data['pack'],
                stock=stock,
                qty=qty,
                gst=gst,
                cgst=cgst,
                sgst=sgst,
                hsn=data['hsn'],
                amount=amount,
                total_amount=total_amount,
                user=request.user
            )

            logger.info(f"ConsumptionItem created: ID={consumption.id}")
            return JsonResponse({
                'redirect_url': '/consumption_list/',
                'status': 'success',
                'message': 'Item consumed and recorded successfully.',
                'consumption_id': consumption.id
            })

    except json.JSONDecodeError:
        logger.warning("Invalid JSON received.")
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON format.'}, status=400)

    except Exception as e:
        logger.exception("Error during consumption creation.")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


import zipfile
def download_database(request):
    db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
    media_path = os.path.join(settings.MEDIA_ROOT)

    # Create an in-memory ZIP file
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        # Add the SQLite3 database to the ZIP file
        zip_file.write(db_path, 'db.sqlite3')

        # Add the media folder to the ZIP file
        for root, dirs, files in os.walk(media_path):
            for file in files:
                file_path = os.path.join(root, file)
                relative_path = os.path.relpath(file_path, media_path)
                zip_file.write(file_path, os.path.join('media', relative_path))

    zip_buffer.seek(0)

    # Serve the ZIP file as a download
    response = FileResponse(zip_buffer, as_attachment=True, filename='Purchase_Software_DB.zip')
    log_usage(
        request.user,
        'BACKUP_DOWNLOAD'
    )
    return response


from django.shortcuts import render
from django.db.models import Sum

from datetime import date
from django.shortcuts import render
from django.db.models import Sum

from django.contrib.auth.models import User

def admin_purchase_dashboard(request):
    items = PurchaseOrderLine.objects.select_related('user', 'vendor', 'item').all()

    totals = items.aggregate(
        total_qty=Sum('qty'),
        total_amount=Sum('total_amount'),
        total_gst=Sum('gst'),
        total_cgst=Sum('cgst'),
        total_sgst=Sum('sgst'),
    )

    # Group by user
    user_totals = (
        items.values('user__username')
        .annotate(user_total=Sum('total_amount'))
        .order_by('-user_total')
    )

    context = {
        'totals': totals,
        'items': items[:50],  # Recent items
        'user_totals': user_totals,
    }

    return render(request, 'admin_purchase_dashboard.html', context)



from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import ConsumedPurchaseOrder, Vendor, Branch, Registration
from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Vendor, Branch, Item, ConsumedPurchaseOrder, Registration


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from datetime import date, timedelta

@login_required
def view_consumption(request):
    # Logged-in user session registration
    bok = get_object_or_404(Registration, id=request.session.get('logg'))

    # Get filters
    vendor_name = request.GET.get('vendor', '').strip()
    item_name = request.GET.get('item', '').strip()
    filter_type = request.GET.get('filter', '').lower()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Pagination parameters
    page = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 10)  # default 10 entries per page

    # Base queryset
    consumed_orders = ConsumedPurchaseOrder.objects.filter(user=request.user)

    # Vendor filter
    if vendor_name:
        consumed_orders = consumed_orders.filter(vendor__name__icontains=vendor_name)

    # Item filter (via related consumed_lines)
    if item_name:
        consumed_orders = consumed_orders.filter(consumed_lines__item__name__icontains=item_name)

    # Date filter setup
    today = date.today()
    if filter_type == 'daily':
        consumed_orders = consumed_orders.filter(dateField=today)
    elif filter_type == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        consumed_orders = consumed_orders.filter(dateField__range=[week_start, week_end])
    elif filter_type == 'monthly':
        consumed_orders = consumed_orders.filter(dateField__year=today.year, dateField__month=today.month)
    elif filter_type == 'yearly':
        consumed_orders = consumed_orders.filter(dateField__year=today.year)
    elif filter_type == 'custom' and start_date and end_date:
        consumed_orders = consumed_orders.filter(dateField__range=[start_date, end_date])

    consumed_orders = consumed_orders.distinct()

    # Totals
    total_quantity = 0
    total_price = 0
    for order in consumed_orders:
        for line in order.consumed_lines.all():
            total_quantity += getattr(line, 'quantity', 0)
            total_price += getattr(line, 'total_price', 0)

    # Pagination
    paginator = Paginator(consumed_orders, per_page)
    try:
        paginated_orders = paginator.page(page)
    except PageNotAnInteger:
        paginated_orders = paginator.page(1)
    except EmptyPage:
        paginated_orders = paginator.page(paginator.num_pages)

    # Context
    context = {
        'vendors': Vendor.objects.filter(user=request.user),
        'items': Item.objects.filter(user=request.user),
        'branches': Branch.objects.all(),
        'consumed_orders': paginated_orders,
        'total_quantity': total_quantity,
        'total_price': total_price,
        'bok': bok,
        'per_page': per_page,
        'paginator': paginator,
        'per_page_options': [10, 25, 50, 100],
    }

    return render(request, 'view_consumption.html', context)

from django.shortcuts import render, get_object_or_404
from datetime import date, timedelta
from .models import ConsumedPurchaseOrder, Vendor, Item, Branch, Registration

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from datetime import date, timedelta

from urllib.parse import urlencode

from datetime import date, timedelta
from urllib.parse import urlencode
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.urls import reverse

@login_required
def custom_consumed(request):
    bok = get_object_or_404(Registration, id=request.session.get('logg'))

    vendor_name = request.GET.get('vendor', '').strip()
    item_name = request.GET.get('item', '').strip()
    filter_type = request.GET.get('filter', 'daily').lower()

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    consumed_orders = ConsumedPurchaseOrder.objects.filter(
        user=request.user
    ).prefetch_related(
        "consumed_lines__item__vendor"
    )

    # ================= FILTERS =================

    if vendor_name:
        consumed_orders = consumed_orders.filter(
            vendor__name__icontains=vendor_name
        )

    if item_name:
        consumed_orders = consumed_orders.filter(
            consumed_lines__item__name__icontains=item_name
        )

    today = date.today()

    if filter_type == "daily":
        consumed_orders = consumed_orders.filter(dateField=today)

    elif filter_type == "weekly":
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        consumed_orders = consumed_orders.filter(
            dateField__range=[week_start, week_end]
        )

    elif filter_type == "monthly":
        consumed_orders = consumed_orders.filter(
            dateField__year=today.year,
            dateField__month=today.month
        )

    elif filter_type == "yearly":
        consumed_orders = consumed_orders.filter(
            dateField__year=today.year
        )

    elif filter_type == "custom" and start_date and end_date:
        consumed_orders = consumed_orders.filter(
            dateField__range=[start_date, end_date]
        )

    consumed_orders = consumed_orders.distinct().order_by("-dateField")

    # ✅ Preserve filters
    query_string = urlencode(request.GET)

    return render(request, "custom_consumed.html", {
        "vendors": Vendor.objects.filter(user=request.user),
        "items": Item.objects.filter(user=request.user),
        "branches": Branch.objects.all(),
        "consumed_orders": consumed_orders,
        "filter_type": filter_type,
        "bok": bok,
        "query_string": query_string,
    })


@login_required
@transaction.atomic
def bulk_delete_consumed_orders(request):

    if request.method == "POST":

        selected_ids = request.POST.getlist("selected_orders")

        if selected_ids:
            ConsumedPurchaseOrder.objects.filter(
                id__in=selected_ids,
                user=request.user
            ).delete()

        # ✅ Preserve previous filters
        next_query = request.POST.get("next", "")

        if next_query:
            return redirect(f"{reverse('custom_consumed')}?{next_query}")

    return redirect("custom_consumed")








from django.http import JsonResponse
import json

def toggle_edit(request, id):
    if request.method == "POST":
        if not request.user.is_authenticated:
            return JsonResponse({"error": "Login required"}, status=403)

        data = json.loads(request.body)
        status = data.get("edit_allowed")  # True / False

        user = Registration.objects.get(id=id)

        # ✅ Convert to string
        user.edit_allowed = "True" if status  else "False"
        user.save()

        return JsonResponse({
            "success": True,
            "edit_allowed": user.edit_allowed
        })


def payment_page(request):
    user_id = request.session.get('logg')
    user = Registration.objects.get(id=user_id)

    amount = 500
    today = now().date()
    next_due = today + timedelta(days=30)

    return render(request, 'payment_page.html', {
        'amount': amount,
        'today': today,
        'next_due': next_due,
        'RAZORPAY_KEY_ID':settings.RAZORPAY_KEY_ID,
    })

def create_order(request):
    client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

    amount = 50000  # ₹500

    order = client.order.create({
        "amount": amount,
        "currency": "INR",
        "payment_capture": 1
    })

    return JsonResponse({
        "order_id": order['id'],
        "amount": amount
    })


from datetime import timedelta
from django.utils.timezone import now

@csrf_exempt
def payment_success(request):
    data = json.loads(request.body)

    user_id = request.session.get('logg')
    user = Registration.objects.get(id=user_id)

    user.payment_status = True
    user.payment_expiry = now() + timedelta(days=30)

    user.save()

    return JsonResponse({"status": "success"})


def billing_activity_list(request):

    activities = BillingActivity.objects.all().order_by('-id')

    return render(
        request,
        'billing/activity_list.html',
        {
            'activities': activities
        }
    )


def billing_activity_add(request):

    if request.method == "POST":

        BillingActivity.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            charge=request.POST.get('charge'),
            description=request.POST.get('description')
        )

        messages.success(
            request,
            "Activity Added Successfully"
        )

        return redirect('billing_activity_list')

    return render(
        request,
        'billing/activity_add.html'
    )


def billing_activity_edit(request, id):

    activity = get_object_or_404(
        BillingActivity,
        id=id
    )

    if request.method == "POST":

        activity.name = request.POST.get('name')
        activity.code = request.POST.get('code')
        activity.charge = request.POST.get('charge')
        activity.description = request.POST.get('description')

        activity.save()

        messages.success(
            request,
            "Activity Updated Successfully"
        )

        return redirect(
            'billing_activity_list'
        )

    return render(
        request,
        'billing/activity_edit.html',
        {
            'activity': activity
        }
    )


def billing_activity_delete(request, id):

    activity = get_object_or_404(
        BillingActivity,
        id=id
    )

    activity.delete()

    messages.success(
        request,
        "Activity Deleted Successfully"
    )

    return redirect(
        'billing_activity_list'
    )


def usage_logs(request):

    logs = UsageLog.objects.select_related(
        'user',
        'activity'
    ).order_by('-created_at')

    return render(
        request,
        'billing/usage_logs.html',
        {
            'logs': logs
        }
    )


from .models import BillingActivity
from .models import UsageLog


def log_usage(user, activity_code, qty=1):

    try:

        activity = BillingActivity.objects.get(
            code=activity_code,
            is_active=True
        )

        UsageLog.objects.create(
            user=user,
            activity=activity,
            quantity=qty,
            charge_per_unit=activity.charge,
            total_charge=activity.charge * qty
        )

    except BillingActivity.DoesNotExist:
        pass


def branch_billing_summary(request):

    branches = User.objects.filter(
        registration__User_role='employee'
    )

    branch_data = []

    for branch in branches:

        total_due = UsageLog.objects.filter(
            user=branch
        ).aggregate(
            total=Sum('total_charge')
        )['total'] or 0

        branch_data.append({
            'branch': branch,
            'total_due': total_due
        })
    print(branch_data,"branchhhh")

    return render(
        request,
        'billing/branch_billing_summary.html',
        {
            'branch_data': branch_data
        }
    )

def branch_billing_detail(request, user_id):

    branch = User.objects.get(id=user_id)

    activity_summary = (
        UsageLog.objects
        .filter(user=branch)
        .values(
            'activity__name'
        )
        .annotate(
            total_count=Count('id'),
            total_amount=Sum('total_charge')
        )
    )

    total_due = UsageLog.objects.filter(
        user=branch
    ).aggregate(
        total=Sum('total_charge')
    )['total'] or 0

    return render(
        request,
        'billing/branch_billing_detail.html',
        {
            'branch': branch,
            'activity_summary': activity_summary,
            'total_due': total_due
        }
    )


def generate_monthly_invoices(request):

    month = datetime.now().month
    year = datetime.now().year

    branches = User.objects.filter(
        registration__User_role='employee'
    )

    for branch in branches:

        total = UsageLog.objects.filter(
            user=branch,
            created_at__month=month,
            created_at__year=year
        ).aggregate(
            total=Sum('total_charge')
        )['total'] or 0

        MonthlyInvoice.objects.get_or_create(
            user=branch,
            month=month,
            year=year,
            defaults={
                'total_amount': total
            }
        )

    messages.success(
        request,
        "Monthly invoices generated successfully."
    )

    return redirect(
        'monthly_invoice_list'
    )

def monthly_invoice_list(request):

    invoices = MonthlyInvoice.objects.select_related(
        'user'
    ).order_by(
        '-year',
        '-month'
    )
    print(invoices,"invv")

    return render(
        request,
        'billing/monthly_invoice_list.html',
        {
            'invoices': invoices
        }
    )


def invoice_payment(request, id):

    invoice = get_object_or_404(
        MonthlyInvoice,
        id=id
    )

    if request.method == "POST":

        invoice.transaction_id = request.POST.get(
            'transaction_id'
        )

        invoice.payment_date = request.POST.get(
            'payment_date'
        )

        invoice.remarks = request.POST.get(
            'remarks'
        )

        invoice.status = request.POST.get(
            'status'
        )

        invoice.save()

        messages.success(
            request,
            "Payment updated successfully."
        )

        return redirect(
            'monthly_invoice_list'
        )

    return render(
        request,
        'billing/invoice_payment.html',
        {
            'invoice': invoice
        }
    )